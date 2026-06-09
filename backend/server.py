from fastapi import FastAPI, APIRouter, File, UploadFile, Form, HTTPException, BackgroundTasks, Request, Depends, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import hashlib
import secrets
import re
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone, timedelta
import base64
# emergentintegrations only works in Emergent environment - using fallback for Railway
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
    EMERGENT_AVAILABLE = True
except ImportError:
    EMERGENT_AVAILABLE = False
    LlmChat = None
    UserMessage = None
    ImageContent = None
import json
import bcrypt
import jwt  # Para decodificar tokens en endpoints de export
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import asyncio
import xlsxwriter
from io import BytesIO
import resend  # Resend como alternativa a SendGrid

# Servicios de seguridad
from services.jwt_service import (
    create_access_token,
    create_refresh_token,
    verify_access_token,
    verify_refresh_token,
    get_current_user,
    require_auth,
    require_admin,
    ADMIN_ROLE_FLAGS,
    JWT_SECRET,
    JWT_ALGORITHM,
    get_token_from_request,
    security
)
from services.rate_limiter import limiter, get_limit, rate_limit_exceeded_handler
from services.global_rate_limiter import GlobalRateLimitMiddleware
from services.audit_service import audit, AuditAction

# Routers modulares
from routes import (
    ia_lab_router,
    auth_advanced_router,
    despiece_budgeter_router,
    libraries_router,
    montajes_router,
    backup_router,
    armarios_router,
    digitalizador_router,
    crm_module_router,
    orders_router,
    ai_engine_router,
)
from routes.clients import router as clients_router
from routes.auth_routes import router as auth_routes_router
from routes.products import router as products_router
from routes.projects import router as projects_router
from routes.fabrica import router as fabrica_router
from routes.dashboard import router as dashboard_router
from routes.factory_reports import router as factory_reports_router
from routes.invoices import router as invoices_router
from routes.command_center import router as command_center_router
from routes.backup import scheduler as backup_scheduler, start_backup_scheduler
from routes.admin import router as admin_router
from routes.rentabilidad import router as rentabilidad_router
from routes.gastos import router as gastos_router
from routes.reports import router as reports_router
from routes.exports import router as exports_router
from routes.maintenance import router as maintenance_router
from routes.telemetry import router as telemetry_router
from routes.users import router as users_router
from routes.settings import router as settings_router
from routes.materials import router as materials_router
from routes.expedient import router as expedient_router
from routes.shop_clients import router as shop_clients_router
from routes.google_calendar import google_calendar_router

# Servicios de backup y tracking
from services.backup_service import init_backup_service
from services.activity_tracker import init_activity_tracker, get_tracker, ActivityType

# Modelos compartidos
from models.schemas import (
    StatusCheck, StatusCheckCreate,
    UserModelInternal, UserResponse, UserCreate, UserUpdate,
    ZonePoints, ProductModel, ProductCreate,
    MaterialModel, MaterialCreate,
    BudgetItemModel, ProjectModel, ProjectCreate, ProjectUpdate,
    SettingsModel, SettingsUpdate,
    ClientModel, ClientCreate, ClientUpdate,
    ContactModel, ContactCreate, ContactUpdate,
    OpportunityModel, OpportunityCreate, OpportunityUpdate,
    CalendarEventModel, CalendarEventCreate, CalendarEventUpdate,
    ActivityModel, ActivityCreate, ActivityUpdate,
    DistributorRequest,
    # Despiece
    DespieceItemInput, DespieceRequest, ComponentPiece, FurnitureDespiece, DespieceResponse,
    # Digitalizador
    DigitalizadorMatchedProduct, DigitalizadorLine, DigitalizadorRequest, DigitalizadorResponse,
    DigitalizadorExportRequest, DigitalizadorSaveRequest, DigitalizadorHistoryItem,
    ExpedienteRequest, DigitalizadorToProjectRequest,
    # Armarios
    ArmarioModuleConfig, ArmarioProject, ArmarioProjectCreate, ArmarioProjectUpdate,
    IAConfigRequest, IARenderRequest, IALayoutRequest,
    # Maintenance
    MaintenanceActivateRequest, MaintenanceStatusResponse,
    # Montadores/Instaladores
    MontadorCreate, MontadorUpdate, MontadorResponse,
    MontajeCreate, MontajeUpdate, MontajeResponse
)


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Create the main app without a prefix
app = FastAPI()

# Configurar Rate Limiter
app.state.limiter = limiter
app.add_exception_handler(429, rate_limit_exceeded_handler)

# ============================================================
# CAPTURA DE ERRORES DE GUARDADO (diagnóstico)
# Registra en db.error_log los fallos de validación (422) y de servidor (500)
# con la ruta, el método, el error y un extracto del cuerpo. Permite ver el
# error exacto sin necesidad de la consola del navegador, vía
# GET /api/admin/recent-errors.
# ============================================================
from fastapi.exceptions import RequestValidationError as _RVE
from fastapi.responses import JSONResponse as _JSONResponse
from starlette.exceptions import HTTPException as _StarletteHTTPException

async def _log_request_error(request, status_code, detail, body_text=None):
    try:
        from datetime import datetime as _dt, timezone as _tz
        doc = {
            "ts": _dt.now(_tz.utc).isoformat(),
            "method": request.method,
            "path": str(request.url.path),
            "status": status_code,
            "detail": detail,
            "body": (body_text or "")[:2000],
        }
        # Solo registrar escrituras (POST/PUT/PATCH/DELETE) para no llenar.
        if request.method in ("POST", "PUT", "PATCH", "DELETE"):
            await db.error_log.insert_one(doc)
    except Exception as _e:
        logger.error(f"No se pudo registrar error de request: {_e}")

@app.exception_handler(_RVE)
async def _validation_error_handler(request, exc):
    body = ""
    try:
        raw = await request.body()
        body = raw.decode("utf-8", "ignore")
    except Exception:
        pass
    # Convertir errores a texto plano (campo + mensaje) para guardar/mostrar.
    errors = [
        {"campo": ".".join(str(p) for p in e.get("loc", [])), "msg": e.get("msg", "")}
        for e in exc.errors()
    ]
    await _log_request_error(request, 422, errors, body)
    return _JSONResponse(status_code=422, content={"detail": errors})

@app.exception_handler(500)
async def _server_error_handler(request, exc):
    await _log_request_error(request, 500, str(exc))
    return _JSONResponse(status_code=500, content={"detail": "Internal Server Error"})

@app.exception_handler(_StarletteHTTPException)
async def _http_exception_handler(request, exc):
    # Captura los errores HTTP lanzados explícitamente (HTTPException 400/401/
    # 403/404/500...), que NO pasan por el handler de 500. Registra los de
    # escritura con código >= 400 y devuelve la respuesta estándar.
    try:
        if exc.status_code >= 400 and request.method in ("POST", "PUT", "PATCH", "DELETE"):
            body = ""
            try:
                raw = await request.body()
                body = raw.decode("utf-8", "ignore")
            except Exception:
                pass
            await _log_request_error(request, exc.status_code, exc.detail, body)
    except Exception:
        pass
    return _JSONResponse(status_code=exc.status_code, content={"detail": exc.detail},
                         headers=getattr(exc, "headers", None))

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Registrar routers modulares
api_router.include_router(ia_lab_router)
api_router.include_router(auth_advanced_router)
api_router.include_router(despiece_budgeter_router)
api_router.include_router(libraries_router)
api_router.include_router(fabrica_router)
api_router.include_router(montajes_router)
api_router.include_router(backup_router)
api_router.include_router(armarios_router)
api_router.include_router(digitalizador_router)
api_router.include_router(crm_module_router)
api_router.include_router(orders_router)
api_router.include_router(dashboard_router)
api_router.include_router(factory_reports_router)
api_router.include_router(invoices_router)
api_router.include_router(command_center_router)
api_router.include_router(admin_router)
api_router.include_router(rentabilidad_router)
api_router.include_router(gastos_router)
api_router.include_router(reports_router)
api_router.include_router(exports_router)
api_router.include_router(maintenance_router)
api_router.include_router(telemetry_router)
api_router.include_router(users_router)
api_router.include_router(settings_router)
api_router.include_router(materials_router)
api_router.include_router(expedient_router)
api_router.include_router(shop_clients_router)
api_router.include_router(ai_engine_router)
api_router.include_router(google_calendar_router)
api_router.include_router(clients_router)
api_router.include_router(auth_routes_router)
api_router.include_router(products_router)
api_router.include_router(projects_router)
# Nota: auth, products, clients, projects están en server.py
# Se integrarán gradualmente para evitar conflictos

# ============================================
# MAINTENANCE MODE STATE
# ============================================
# Estado del modo mantenimiento (en memoria, se sincroniza con BD)
maintenance_state = {
    "active": False,
    "message": "Sistema en actualización. Volvemos pronto.",
    "activatedAt": None,
    "activatedBy": None,
    "estimatedEndTime": None,
    "preUpdateBackupId": None
}


# ============================================
# PASSWORD UTILITIES
# ============================================

def hash_password(password: str) -> str:
    """Hash a password using bcrypt"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

# verify_password movido a routes/auth_routes.py


# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Hello World"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    
    # Convert to dict and serialize datetime to ISO string for MongoDB
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    
    _ = await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    # Exclude MongoDB's _id field from the query results
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime objects
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    
    return status_checks


# =============================================
# SISTEMA DE COLA DE TELEMETRÍA IA
# =============================================
from services.telemetry_queue import (
    create_telemetry_job, 
    get_job_status,
    get_audit_images,
    get_audit_image_detail,
    update_audit_notes
)

@api_router.post("/analyze-product-sheets")
async def analyze_product_sheets(
    module: str = Form(...),
    library: str = Form("ZC"),
    files: List[UploadFile] = File(...)
):
    """
    Analiza fichas de productos usando Gemini Vision API.
    Extrae: código, nombre, dimensiones, puntos por zona, categoría, serie.
    Detecta AUTOMÁTICAMENTE nuevas categorías desde el encabezado de la página.
    """
    try:
        api_key = os.environ.get('EMERGENT_LLM_KEY')
        if not api_key:
            return {"error": "EMERGENT_LLM_KEY not configured"}
        
        # Log library being used
        logger.info(f"Analyzing product sheets for library: {library}")
        
        products = []
        detected_categories = set()
        skipped_files = []
        
        for file in files:
            # Read file content
            content = await file.read()
            
            # Determine file type and handle accordingly
            filename = file.filename.lower() if file.filename else ""
            content_type = file.content_type or ""
            
            # Check if it's a PDF - Gemini doesn't support PDFs directly
            if filename.endswith('.pdf') or 'pdf' in content_type.lower():
                logger.warning(f"PDF file detected: {file.filename}. PDFs are not directly supported by Gemini Vision. Please convert to JPG/PNG first.")
                skipped_files.append({"filename": file.filename, "reason": "PDFs no soportados. Convierte a JPG/PNG primero."})
                continue  # Skip PDF files for now
            
            # Validate image format
            valid_formats = ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp']
            is_valid_image = any(filename.endswith(fmt) for fmt in valid_formats)
            
            if not is_valid_image:
                # Try to detect from content type
                valid_mime_types = ['image/jpeg', 'image/png', 'image/webp', 'image/gif', 'image/bmp']
                if content_type.lower() not in valid_mime_types:
                    logger.warning(f"Unsupported file format: {file.filename} ({content_type}). Skipping.")
                    skipped_files.append({"filename": file.filename, "reason": f"Formato no soportado: {content_type}"})
                    continue
            
            # Validate that we have actual image data
            if len(content) < 100:
                logger.warning(f"File {file.filename} is too small to be a valid image. Skipping.")
                skipped_files.append({"filename": file.filename, "reason": "Archivo demasiado pequeño"})
                continue
            
            # Check magic bytes for common image formats
            magic_bytes = content[:8]
            is_jpeg = magic_bytes[:2] == b'\xff\xd8'
            is_png = magic_bytes[:8] == b'\x89PNG\r\n\x1a\n'
            is_gif = magic_bytes[:6] in (b'GIF87a', b'GIF89a')
            is_webp = magic_bytes[8:12] == b'WEBP' if len(content) > 12 else False
            
            if not (is_jpeg or is_png or is_gif or is_webp):
                # Check if it's actually a PDF masquerading as an image
                if magic_bytes[:4] == b'%PDF':
                    logger.warning(f"File {file.filename} is actually a PDF. PDFs are not supported. Please convert to JPG/PNG.")
                    skipped_files.append({"filename": file.filename, "reason": "Es un PDF. Convierte a JPG/PNG."})
                    continue
                logger.warning(f"File {file.filename} does not appear to be a valid image format. Attempting anyway...")
            
            base64_image = base64.b64encode(content).decode('utf-8')
            
            # Para MV, hacer múltiples pasadas para extraer todos los productos
            if library == 'MV':
                file_products = []
                extracted_codes = set()
                max_passes = 2  # 2 pasadas = hasta 100 productos por imagen (más rápido)
                detected_tariff = None  # Tarifa detectada automáticamente
                
                for pass_num in range(max_passes):
                    # Construir lista de códigos ya extraídos para excluirlos
                    exclude_list = ", ".join(list(extracted_codes)[:50]) if extracted_codes else "ninguno"
                    
                    system_prompt = f"""Eres un experto en digitalización de tarifas de muebles MV (MUEBLES VALENCIA).
Tu tarea es extraer productos de la tabla de tarifas.

═══════════════════════════════════════════════════════════════
PASO 0: DETECTAR NÚMERO DE TARIFA DESDE EL ENCABEZADO
═══════════════════════════════════════════════════════════════
IMPORTANTE: Lee el encabezado/título de la página para identificar el NÚMERO DE TARIFA.
Busca textos como:
- "TARIFA 1", "TARIFA 2", "TARIFA 3" ... "TARIFA 21"
- "T1", "T2", "T3" ... "T21"
- "PRECIO TARIFA 1", "PVP TARIFA 3", etc.

El número de tarifa es CRÍTICO para mapear correctamente los precios.
Si no encuentras un número de tarifa explícito, usa "T1" por defecto.

═══════════════════════════════════════════════════════════════
ESTRUCTURA DE TARIFAS MV:
═══════════════════════════════════════════════════════════════
- Las tablas tienen secciones: ALTO, ALTO CAMPANA, ALTO RINCON, ALTO VITRINA, BAJO, SEMICOLUMNA, COLUMNA, PUERTAS, etc.
- Cada producto tiene un código (A25D/I*, ASCE60D/I*, B30D/I*, etc.)
- Las columnas numéricas (70, 90, 127, 147, etc.) son ALTURAS en cm
- Los valores en las celdas son los PRECIOS

CÓDIGOS DE PRODUCTO MV:
- A = Alto (A25D/I*, A30D/I*, etc.)
- ASCE = Alto Campana Esquina
- AR = Alto Rincon
- AV = Alto Vitrina
- AD = Alto Decorativo
- AE = Alto Escurreplatos
- AM = Alto Microondas
- B = Bajo
- SC = Semicolumna
- C = Columna
- P = Puerta

FORMATO DE SALIDA (JSON):
{{
  "detectedTariff": "T3",
  "products": [
    {{
      "code": "A30D/I*-70",
      "name": "Alto 30 D/I 70cm",
      "category": "ALTO",
      "width": 300,
      "height": 700,
      "points": 45
    }}
  ]
}}

REGLAS:
1. Código = código_producto + "-" + altura (ej: "A30D/I*-70", "ASCE60D/I*-90")
2. points = precio de la celda
3. width = número en el código (A30 = 300mm, A60 = 600mm)
4. height = columna de altura en mm (70 = 700mm, 90 = 900mm)
5. "detectedTariff" debe ser el número de tarifa detectado del encabezado (T1, T2, T3... T21)

{"IMPORTANTE: NO incluyas estos códigos que ya fueron extraídos: " + exclude_list if extracted_codes else ""}

Extrae hasta 50 productos. Responde SOLO con JSON válido."""

                    chat = LlmChat(
                        api_key=api_key,
                        session_id=f"product-analysis-mv-{uuid.uuid4()}",
                        system_message=system_prompt
                    ).with_model("gemini", "gemini-2.5-flash")
                    
                    user_prompt = f"""Analiza esta página de TARIFA MV.
PRIMERO: Detecta el número de tarifa del encabezado (T1, T2, T3... T21).
{"SEGUNDO: Extrae productos DIFERENTES a los ya extraídos." if extracted_codes else "SEGUNDO: Extrae todos los productos que veas."}
Responde SOLO con JSON válido con "detectedTariff" y "products"."""
                    
                    user_message = UserMessage(
                        text=user_prompt,
                        file_contents=[ImageContent(image_base64=base64_image)]
                    )
                    
                    try:
                        response = await chat.send_message(user_message)
                        response_text = str(response).strip() if response else ""
                        
                        # Limpiar markdown
                        if response_text.startswith("```"):
                            response_text = response_text.split("```")[1]
                            if response_text.startswith("json"):
                                response_text = response_text[4:]
                        response_text = response_text.strip()
                        
                        parsed = json.loads(response_text)
                        new_products = parsed.get("products", [])
                        
                        # Detectar tarifa del primer pase
                        if pass_num == 0:
                            detected_tariff = parsed.get("detectedTariff", "T1")
                            # Validar formato de tarifa
                            if not detected_tariff or not detected_tariff.startswith("T"):
                                detected_tariff = "T1"
                            logger.info(f"MV Tarifa detectada automáticamente: {detected_tariff}")
                        
                        # Filtrar productos nuevos (no duplicados)
                        new_count = 0
                        for prod in new_products:
                            code = prod.get("code", "")
                            if code and code not in extracted_codes:
                                prod['id'] = f"AI-{module.upper()}-{uuid.uuid4().hex[:8]}"
                                prod['manufacturer'] = 'MV'
                                prod['module'] = module
                                prod['library'] = library
                                prod['importedAt'] = datetime.now(timezone.utc).isoformat()
                                prod['originalFilename'] = file.filename
                                # Usar tarifa detectada para zonePoints
                                prod['zonePoints'] = {detected_tariff: prod.get('points', 0)}
                                prod['detectedTariff'] = detected_tariff
                                file_products.append(prod)
                                extracted_codes.add(code)
                                new_count += 1
                        
                        logger.info(f"MV Pass {pass_num+1}: {new_count} nuevos productos (total: {len(file_products)}, tarifa: {detected_tariff})")
                        
                        # Si no encontró productos nuevos, terminar
                        if new_count == 0:
                            break
                            
                    except Exception as pass_error:
                        logger.warning(f"MV Pass {pass_num+1} error: {pass_error}")
                        break
                
                products.extend(file_products)
                logger.info(f"MV Total from {file.filename}: {len(file_products)} productos (tarifa detectada: {detected_tariff})")
                continue  # Saltar el procesamiento de ZC para este archivo
                
            else:
                # Código original para ZC
                system_prompt = """Eres un experto en digitalización de tarifas técnicas de muebles de cocina ZONA COCINAS.
Tu tarea es extraer TODOS los productos visibles en la imagen de forma estructurada.

═══════════════════════════════════════════════════════════════
PASO 1: DETECTAR CATEGORÍA PRINCIPAL DESDE EL ENCABEZADO
═══════════════════════════════════════════════════════════════
Lee el título/encabezado de CADA página para identificar la categoría.
El encabezado suele estar en la parte superior con formato:
"PROGRAMA ESTÁNDAR - [TIPO DE MÓDULO] - [SERIE/FONDO]"

CATEGORÍAS PRINCIPALES DEL CATÁLOGO ZONA COCINAS:

📦 MÓDULOS ALTOS (category: "ALTOS"):
   - "ALTOS 35 FONDO 58" → series: "ALTOS 35 F58"
   - "ALTOS 40 FONDO 33" → series: "ALTOS 40 F33"
   - "ALTOS ABATIBLES" → series: "ABATIBLES"
   - "ALTOS ESQUINEROS" → series: "ESQUINEROS"
   - Códigos empiezan por: 35A, 40A, 35AB, etc.

📦 MÓDULOS BAJOS (category: "BAJOS"):
   - "BAJOS 70 FONDO ESTÁNDAR" → series: "BAJOS 70"
   - "BAJOS 80 FONDO ESTÁNDAR" → series: "BAJOS 80"
   - "BAJOS FREGADERO" → series: "FREGADERO"
   - "BAJOS HORNO" → series: "HORNO"
   - "BAJOS ESQUINEROS" → series: "ESQUINEROS"
   - Códigos empiezan por: 7B, 8B, 70B, 80B, etc.

📦 SEMICOLUMNAS (category: "SEMICOLUMNAS"):
   - "SEMICOLUMNAS 135/140" → series: "SC 135-140"
   - "SEMICOLUMNAS 160/165" → series: "SC 160-165"
   - Códigos empiezan por: 135SC, 140SC, 160SC, etc.

📦 COLUMNAS (category: "COLUMNAS"):
   - "COLUMNAS 195/200" → series: "COL 195-200"
   - "COLUMNAS 215/220" → series: "COL 215-220"
   - "COLUMNAS DESPENSA" → series: "DESPENSA"
   - "COLUMNAS HORNO/MICRO" → series: "HORNO-MICRO"
   - Códigos empiezan por: 195C, 200C, 215C, etc.

🚪 PUERTAS Y VITRINAS (category: "PUERTAS"):
   - "PUERTAS" → series: "PUERTAS"
   - "VITRINAS" → series: "VITRINAS"
   - "PUERTAS LACADAS" → series: "LACADAS"
   - Códigos empiezan por: PTA_, VIT_, etc.

🔧 ACCESORIOS (category: "ACCESORIOS"):
   - "ACCESORIOS" → series: "ACCESORIOS"
   - "HERRAJES" → series: "HERRAJES"
   - "TIRADORES" → series: "TIRADORES"

📐 OTROS ELEMENTOS:
   - "ZÓCALOS Y CORNISAS" → category: "ZOCALOS"
   - "ENCIMERAS" → category: "ENCIMERAS"
   - "ELECTRODOMÉSTICOS" → category: "ELECTRO"
   - "COMPLEMENTOS" → category: "COMPLEMENTOS"

═══════════════════════════════════════════════════════════════
PASO 2: EXTRAER CÓDIGOS Y PRECIOS
═══════════════════════════════════════════════════════════════
Para CADA fila de la tabla:
1. Lee el CÓDIGO exacto (primera columna, ej: 35A1P58350)
2. Lee los 12 valores de precio Z1 a Z12 (de izquierda a derecha)

DECODIFICACIÓN DE CÓDIGOS:
- 35A1P58350: 35=altura(cm), A=alto, 1P=1puerta, 58=fondo(cm), 350=ancho(mm)
- 35A2P58600: 35=altura, A=alto, 2P=2puertas, 58=fondo, 600=ancho
- 7B1P300: 7=70cm altura, B=bajo, 1P=1puerta, 300=ancho(mm)
- 80B2P600: 80=80cm altura, B=bajo, 2P=2puertas, 600=ancho
- 135SC1P58450: 135=altura, SC=semicolumna, 1P=1puerta, 58=fondo, 450=ancho
- 200C2P60600: 200=altura, C=columna, 2P=2puertas, 60=fondo, 600=ancho
- PTA_ZC898X298: PTA=puerta, ZC=serie, 898=ancho(mm), 298=alto(mm)

═══════════════════════════════════════════════════════════════
FORMATO DE RESPUESTA JSON:
═══════════════════════════════════════════════════════════════
{
  "detectedCategory": "ALTOS|BAJOS|SEMICOLUMNAS|COLUMNAS|PUERTAS|ACCESORIOS|etc",
  "detectedSubcategory": "Serie o tipo específico del encabezado",
  "pageTitle": "Texto exacto del encabezado de la página",
  "products": [
    {
      "code": "CÓDIGO_EXACTO",
      "name": "Descripción: [Tipo] [Altura]cm [NºPuertas] [Ancho]mm",
      "category": "CATEGORÍA",
      "series": "Serie del encabezado",
      "visualType": "1P/2P/ABATIBLE/VITRINA/FREGADERO/etc",
      "width": ancho_mm,
      "height": altura_cm,
      "depth": fondo_cm,
      "points": valor_Z1,
      "zonePoints": {"Z1":n,"Z2":n,"Z3":n,"Z4":n,"Z5":n,"Z6":n,"Z7":n,"Z8":n,"Z9":n,"Z10":n,"Z11":n,"Z12":n}
    }
  ]
}

REGLAS CRÍTICAS:
✅ Lee el encabezado EXACTO de la página para determinar categoría
✅ Extrae TODOS los productos de la tabla, no solo algunos
✅ Los códigos deben ser EXACTOS como aparecen
✅ Z1 es el PRIMER precio de cada fila (columna después del código)
✅ Responde SOLO con JSON válido, sin explicaciones"""
            
            # Create Gemini chat with vision
            chat = LlmChat(
                api_key=api_key,
                session_id=f"product-analysis-{uuid.uuid4()}",
                system_message=system_prompt
            ).with_model("gemini", "gemini-2.5-flash")
            
            # Create message with image - different prompt for MV vs ZC
            if library == 'MV':
                user_prompt = """Analiza esta página de TARIFA MV.
Extrae los productos de la tabla siguiendo el formato JSON especificado.
MÁXIMO 30 productos. Si hay más, extrae solo los primeros 30.
Responde SOLO con JSON válido."""
            else:
                user_prompt = """Analiza esta página de tarifa técnica de ZONA COCINAS.

INSTRUCCIONES:
1. Lee el ENCABEZADO de la página para identificar: ALTOS, BAJOS, SEMICOLUMNAS, COLUMNAS, PUERTAS, ACCESORIOS, etc.
2. Identifica la SERIE (ej: "ALTOS 35 FONDO 58", "BAJOS 70", "COLUMNAS 200")
3. Extrae TODOS los códigos de productos de la tabla (35A1P58350, 7B1P300, etc.)
4. Para cada producto, lee los 12 precios por zona (Z1 a Z12)
5. Decodifica las dimensiones del código

Responde ÚNICAMENTE con el JSON estructurado. No añadas explicaciones."""
            
            # Create the user message with image
            user_message = UserMessage(
                text=user_prompt,
                file_contents=[ImageContent(image_base64=base64_image)]
            )
            
            # Get AI response
            response = await chat.send_message(user_message)
            
            # Parse JSON response
            try:
                # Clean response (remove markdown code blocks if present)
                clean_response = response.strip()
                
                # Handle markdown code blocks more robustly
                if '```json' in clean_response:
                    # Extract content between ```json and ```
                    start = clean_response.find('```json') + 7
                    end = clean_response.find('```', start)
                    if end == -1:
                        # No closing ```, take everything after ```json
                        clean_response = clean_response[start:].strip()
                    else:
                        clean_response = clean_response[start:end].strip()
                elif '```' in clean_response:
                    # Generic code block
                    start = clean_response.find('```') + 3
                    end = clean_response.find('```', start)
                    if end == -1:
                        clean_response = clean_response[start:].strip()
                    else:
                        clean_response = clean_response[start:end].strip()
                
                # Try to find JSON object or array
                import re
                if not clean_response.startswith('{') and not clean_response.startswith('['):
                    json_match = re.search(r'(\{[\s\S]*\}|\[[\s\S]*\])', clean_response)
                    if json_match:
                        clean_response = json_match.group(1)
                
                parsed_data = json.loads(clean_response)
                
                # Handle new format with detectedCategory
                if isinstance(parsed_data, dict) and 'products' in parsed_data:
                    # New format with category detection
                    file_category = parsed_data.get('detectedCategory', 'OTROS')
                    file_subcategory = parsed_data.get('detectedSubcategory', '')
                    product_list = parsed_data.get('products', [])
                    detected_categories.add(file_category)
                    logger.info(f"Detected category: {file_category}, subcategory: {file_subcategory}")
                elif isinstance(parsed_data, list):
                    # Old format - array of products
                    product_list = parsed_data
                    file_category = None
                    file_subcategory = None
                else:
                    # Single product
                    product_list = [parsed_data]
                    file_category = parsed_data.get('category', 'OTROS')
                    file_subcategory = None
                
                for product_data in product_list:
                    # Add metadata
                    product_data['id'] = f"AI-{module.upper()}-{uuid.uuid4().hex[:8]}"
                    product_data['manufacturer'] = 'Zona Cocinas'
                    product_data['module'] = module
                    product_data['library'] = library  # Usar la biblioteca seleccionada (MV o ZC)
                    product_data['importedAt'] = datetime.now(timezone.utc).isoformat()
                    product_data['originalFilename'] = file.filename
                    
                    # Apply detected category if product doesn't have one
                    if file_category and not product_data.get('category'):
                        product_data['category'] = file_category
                    if file_subcategory and not product_data.get('series'):
                        product_data['series'] = file_subcategory
                    
                    products.append(product_data)
                
            except json.JSONDecodeError as e:
                # Try to repair truncated JSON
                logger.warning(f"JSON parse error, attempting repair: {e}")
                try:
                    # Try to extract complete products from truncated response
                    import re
                    # Find all complete product objects
                    product_pattern = r'\{\s*"code":\s*"[^"]+",\s*"name":\s*"[^"]*",\s*"category":\s*"[^"]*"[^}]*"zonePoints":\s*\{[^}]+\}\s*\}'
                    found_products = re.findall(product_pattern, clean_response)
                    
                    if found_products:
                        logger.info(f"Recovered {len(found_products)} products from truncated response")
                        # Extract category from response
                        cat_match = re.search(r'"detectedCategory":\s*"([^"]+)"', clean_response)
                        file_category = cat_match.group(1) if cat_match else 'OTROS'
                        detected_categories.add(file_category)
                        
                        for prod_json in found_products:
                            try:
                                prod = json.loads(prod_json)
                                prod['id'] = f"AI-{module.upper()}-{uuid.uuid4().hex[:8]}"
                                prod['manufacturer'] = 'Zona Cocinas'
                                prod['module'] = module
                                prod['library'] = library  # Usar la biblioteca seleccionada (MV o ZC)
                                prod['importedAt'] = datetime.now(timezone.utc).isoformat()
                                prod['originalFilename'] = file.filename
                                if not prod.get('category'):
                                    prod['category'] = file_category
                                products.append(prod)
                            except:
                                pass
                    else:
                        # Could not recover
                        logger.error(f"Could not repair JSON: {response[:500]}")
                        products.append({
                            "error": f"No se pudo parsear la respuesta para {file.filename}",
                            "raw_response": response[:500]
                        })
                except Exception as repair_error:
                    logger.error(f"Error repairing JSON: {repair_error}. Original response: {response[:500]}")
                    products.append({
                        "error": f"No se pudo parsear la respuesta para {file.filename}",
                        "raw_response": response[:500]
                    })
        
        return {
            "success": True,
            "count": len(products),
            "products": products,
            "detectedCategories": list(detected_categories),
            "skippedFiles": skipped_files
        }
        
    except Exception as e:
        logger.error(f"Error in analyze_product_sheets: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }

# ============================================
# USER ENDPOINTS
# ============================================

# user_to_response movido a routes/auth_routes.py


# ============================================
# DISTRIBUTOR REQUEST ENDPOINTS
# ============================================

class DistributorRequest(BaseModel):
    """Solicitud de alta de distribuidor"""
    companyName: str
    contactName: str
    email: str
    phone: str
    city: str = ""
    province: str = ""
    message: str = ""


@api_router.post("/distributor/request")
@limiter.limit(get_limit("register"))
async def request_distributor(request: Request, data: DistributorRequest):
    """
    Recibir solicitud de alta de distribuidor y enviar email al administrador
    """
    try:
        # Guardar la solicitud en la base de datos
        request_data = {
            "id": f"dist-req-{uuid.uuid4().hex[:8]}",
            "companyName": data.companyName,
            "contactName": data.contactName,
            "email": data.email,
            "phone": data.phone,
            "city": data.city,
            "province": data.province,
            "message": data.message,
            "status": "pending",  # pending, approved, rejected
            "createdAt": datetime.now(timezone.utc).isoformat(),
            "processedAt": None,
            "processedBy": None,
            "notes": ""
        }
        
        await db.distributor_requests.insert_one(request_data)
        
        # Enviar email al administrador
        admin_email = "mario@luiggihome.es"
        
        email_html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #f8fafc; padding: 20px;">
            <div style="background: linear-gradient(135deg, #1e1b4b 0%, #312e81 100%); color: white; padding: 30px; border-radius: 16px 16px 0 0; text-align: center;">
                <h1 style="margin: 0; font-size: 24px;">LUIGGI HOME</h1>
                <p style="margin: 8px 0 0 0; opacity: 0.8; font-size: 14px;">Nueva Solicitud de Distribuidor</p>
            </div>
            
            <div style="background: white; padding: 30px; border-radius: 0 0 16px 16px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                <div style="background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin-bottom: 20px; border-radius: 0 8px 8px 0;">
                    <p style="margin: 0; color: #92400e; font-weight: bold; font-size: 14px;">
                        ⚡ Nueva solicitud de alta pendiente de revisión
                    </p>
                </div>
                
                <h2 style="color: #1e1b4b; margin: 0 0 20px 0; font-size: 18px; border-bottom: 2px solid #e2e8f0; padding-bottom: 10px;">
                    Datos del Solicitante
                </h2>
                
                <table style="width: 100%; border-collapse: collapse;">
                    <tr>
                        <td style="padding: 12px 0; border-bottom: 1px solid #f1f5f9; color: #64748b; font-size: 13px; width: 35%;">
                            <strong>Empresa:</strong>
                        </td>
                        <td style="padding: 12px 0; border-bottom: 1px solid #f1f5f9; color: #1e293b; font-size: 14px; font-weight: 600;">
                            {data.companyName}
                        </td>
                    </tr>
                    <tr>
                        <td style="padding: 12px 0; border-bottom: 1px solid #f1f5f9; color: #64748b; font-size: 13px;">
                            <strong>Contacto:</strong>
                        </td>
                        <td style="padding: 12px 0; border-bottom: 1px solid #f1f5f9; color: #1e293b; font-size: 14px;">
                            {data.contactName}
                        </td>
                    </tr>
                    <tr>
                        <td style="padding: 12px 0; border-bottom: 1px solid #f1f5f9; color: #64748b; font-size: 13px;">
                            <strong>Email:</strong>
                        </td>
                        <td style="padding: 12px 0; border-bottom: 1px solid #f1f5f9;">
                            <a href="mailto:{data.email}" style="color: #ea580c; text-decoration: none; font-weight: 600;">{data.email}</a>
                        </td>
                    </tr>
                    <tr>
                        <td style="padding: 12px 0; border-bottom: 1px solid #f1f5f9; color: #64748b; font-size: 13px;">
                            <strong>Teléfono:</strong>
                        </td>
                        <td style="padding: 12px 0; border-bottom: 1px solid #f1f5f9;">
                            <a href="tel:{data.phone}" style="color: #ea580c; text-decoration: none; font-weight: 600;">{data.phone}</a>
                        </td>
                    </tr>
                    <tr>
                        <td style="padding: 12px 0; border-bottom: 1px solid #f1f5f9; color: #64748b; font-size: 13px;">
                            <strong>Ubicación:</strong>
                        </td>
                        <td style="padding: 12px 0; border-bottom: 1px solid #f1f5f9; color: #1e293b; font-size: 14px;">
                            {data.city}{', ' + data.province if data.province else ''}{' - Sin especificar' if not data.city else ''}
                        </td>
                    </tr>
                </table>
                
                {f'''
                <div style="margin-top: 20px; padding: 15px; background: #f8fafc; border-radius: 8px;">
                    <p style="margin: 0 0 8px 0; color: #64748b; font-size: 12px; font-weight: bold; text-transform: uppercase;">Mensaje:</p>
                    <p style="margin: 0; color: #475569; font-size: 14px; line-height: 1.6;">{data.message}</p>
                </div>
                ''' if data.message else ''}
                
                <div style="margin-top: 30px; text-align: center;">
                    <p style="color: #94a3b8; font-size: 12px; margin: 0;">
                        Solicitud recibida el {datetime.now(timezone.utc).strftime('%d/%m/%Y a las %H:%M')} UTC
                    </p>
                </div>
            </div>
            
            <div style="text-align: center; padding: 20px; color: #94a3b8; font-size: 11px;">
                <p style="margin: 0;">LUIGGI HOME - Sistema de Presupuestos Profesional</p>
            </div>
        </div>
        """
        
        # Intentar enviar email
        sendgrid_key = os.environ.get('SENDGRID_API_KEY')
        if sendgrid_key:
            try:
                sg = SendGridAPIClient(sendgrid_key)
                message = Mail(
                    from_email=os.environ.get('SENDGRID_FROM_EMAIL', 'noreply@luiggihome.es'),
                    to_emails=admin_email,
                    subject=f"🏢 Nueva Solicitud de Distribuidor: {data.companyName}",
                    html_content=email_html
                )
                sg.send(message)
                logger.info(f"Distributor request email sent to {admin_email}")
            except Exception as e:
                logger.error(f"Error sending distributor request email: {e}")
        else:
            logger.warning("SendGrid not configured - distributor request email not sent")
        
        # Auditoría
        audit.log(
            AuditAction.USER_CREATE,
            resource_type="distributor_request",
            resource_id=request_data["id"],
            request=request,
            details={"company": data.companyName, "email": data.email}
        )
        
        if "_id" in request_data:
            del request_data["_id"]
        
        return {
            "success": True,
            "message": "Solicitud enviada correctamente",
            "requestId": request_data["id"]
        }
        
    except Exception as e:
        logger.error(f"Error processing distributor request: {e}")
        raise HTTPException(status_code=500, detail="Error al procesar la solicitud")


@api_router.get("/distributor/requests")
async def get_distributor_requests(
    status: Optional[str] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Obtener todas las solicitudes de distribuidores (solo admins)"""
    # Verificar que es admin
    if credentials:
        try:
            payload = verify_access_token(credentials.credentials)
            if not payload.get("isAdmin"):
                raise HTTPException(status_code=403, detail="Acceso denegado")
        except:
            raise HTTPException(status_code=401, detail="Token inválido")
    else:
        raise HTTPException(status_code=401, detail="Autenticación requerida")
    
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.distributor_requests.find(query, {"_id": 0}).sort("createdAt", -1).to_list(500)
    return requests


@api_router.put("/distributor/requests/{request_id}")
async def update_distributor_request(
    request_id: str,
    data: dict,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Actualizar estado de una solicitud de distribuidor (solo admins)"""
    # Verificar que es admin
    if credentials:
        try:
            payload = verify_access_token(credentials.credentials)
            if not payload.get("isAdmin"):
                raise HTTPException(status_code=403, detail="Acceso denegado")
            admin_id = payload.get("sub")
            admin_username = payload.get("username")
        except:
            raise HTTPException(status_code=401, detail="Token inválido")
    else:
        raise HTTPException(status_code=401, detail="Autenticación requerida")
    
    existing = await db.distributor_requests.find_one({"id": request_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    
    update_data = {
        "status": data.get("status", existing.get("status")),
        "notes": data.get("notes", existing.get("notes", "")),
        "processedAt": datetime.now(timezone.utc).isoformat(),
        "processedBy": admin_username
    }
    
    await db.distributor_requests.update_one({"id": request_id}, {"$set": update_data})
    
    updated = await db.distributor_requests.find_one({"id": request_id}, {"_id": 0})
    return updated


# ============================================
# CLIENT ENDPOINTS - Clientes Activos
# ============================================

# Los endpoints de CLIENTES viven ahora en routes/clients.py
# (extraídos tal cual; registrados via clients_router más arriba).

# Los endpoints de AUTH viven ahora en routes/auth_routes.py

# Los endpoints de PRODUCTOS viven ahora en routes/products.py

# Los endpoints de PROYECTOS viven ahora en routes/projects.py

@api_router.get("/admin/all-work")
async def get_all_work_for_admin(current_user: dict = Depends(require_admin)):
    """
    [ADMIN ONLY] Obtener todos los trabajos de todos los usuarios.
    Incluye proyectos, oportunidades y digitalizaciones.
    """
    # Get all projects grouped by user
    projects = await db.projects.find({}, {"_id": 0}).to_list(5000)
    
    # Get all opportunities
    opportunities = await db.opportunities.find({}, {"_id": 0}).to_list(5000)
    
    # Get all digitalizador history
    digitalizaciones = await db.digitalizador_history.find({}, {"_id": 0}).to_list(5000)
    
    # Get all users for reference
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(500)
    users_map = {u["id"]: u for u in users}
    
    # Enrich projects with user info
    for proj in projects:
        user = users_map.get(proj.get("userId", ""))
        proj["userName"] = user.get("username", "Desconocido") if user else "Desconocido"
        proj["userClientName"] = user.get("clientName", "") if user else ""
    
    # Enrich opportunities with user info
    for opp in opportunities:
        user = users_map.get(opp.get("assignedTo", ""))
        opp["userName"] = user.get("username", "Desconocido") if user else "Sin asignar"
    
    # Enrich digitalizaciones with user info
    for dig in digitalizaciones:
        user = users_map.get(dig.get("createdBy", ""))
        dig["userName"] = user.get("username", "Desconocido") if user else "Desconocido"
    
    return {
        "projects": projects,
        "opportunities": opportunities,
        "digitalizaciones": digitalizaciones,
        "users": users,
        "summary": {
            "totalProjects": len(projects),
            "totalOpportunities": len(opportunities),
            "totalDigitalizaciones": len(digitalizaciones),
            "totalUsers": len(users)
        }
    }

@api_router.get("/admin/metrics")
async def get_admin_metrics(current_user: dict = Depends(require_admin)):
    """
    [DIRECTOR COMERCIAL ONLY] Métricas completas del sistema por delegación y comercial.
    Incluye ventas, oportunidades, rendimiento de comerciales, etc.
    """
    from datetime import timedelta
    
    # Get all users
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(500)
    users_map = {u["id"]: u for u in users}
    
    # Separar usuarios por rol
    directores = [u for u in users if u.get("isAdmin")]
    responsables = [u for u in users if u.get("isResponsableDelegacion")]
    comerciales = [u for u in users if u.get("isRepresentative")]
    tiendas = [u for u in users if u.get("isTienda")]
    colaboradores = [u for u in users if u.get("isPrescriptor")]
    
    # Get all opportunities
    opportunities = await db.opportunities.find({}, {"_id": 0}).to_list(5000)
    
    # Get all projects
    projects = await db.projects.find({}, {"_id": 0}).to_list(5000)
    
    # Get all contacts
    contacts = await db.contacts.find({}, {"_id": 0}).to_list(5000)
    
    # Calcular métricas por comercial/responsable
    metrics_by_user = {}
    
    for user in comerciales + responsables:
        user_id = user["id"]
        
        # Oportunidades asignadas a este usuario
        user_opps = [o for o in opportunities if o.get("assignedTo") == user_id]
        won_opps = [o for o in user_opps if o.get("stage") == "won"]
        active_opps = [o for o in user_opps if o.get("stage") not in ["won", "lost"]]
        
        # Valor total
        total_value = sum(o.get("value", 0) for o in won_opps)
        pipeline_value = sum(o.get("value", 0) for o in active_opps)
        
        # Contactos asignados
        user_contacts = [c for c in contacts if c.get("assignedTo") == user_id]
        
        # Tiendas bajo su gestión
        user_shops = [t for t in tiendas if t.get("linkedRepresentativeId") == user_id]
        
        # Proyectos de sus tiendas
        shop_ids = [s["id"] for s in user_shops]
        user_projects = [p for p in projects if p.get("userId") in shop_ids or p.get("userId") == user_id]
        
        metrics_by_user[user_id] = {
            "userId": user_id,
            "userName": user.get("clientName", user.get("username", "Desconocido")),
            "role": "Resp. Delegación" if user.get("isResponsableDelegacion") else "Comercial",
            "totalOpportunities": len(user_opps),
            "wonOpportunities": len(won_opps),
            "activeOpportunities": len(active_opps),
            "totalValue": total_value,
            "pipelineValue": pipeline_value,
            "conversionRate": round(len(won_opps) / len(user_opps) * 100, 1) if user_opps else 0,
            "totalContacts": len(user_contacts),
            "totalShops": len(user_shops),
            "totalProjects": len(user_projects),
            "shops": [{"id": s["id"], "name": s.get("clientName", s.get("username"))} for s in user_shops]
        }
    
    # Métricas globales
    all_won = [o for o in opportunities if o.get("stage") == "won"]
    all_active = [o for o in opportunities if o.get("stage") not in ["won", "lost"]]
    
    global_metrics = {
        "totalUsers": len(users),
        "totalDirectores": len(directores),
        "totalResponsables": len(responsables),
        "totalComerciales": len(comerciales),
        "totalTiendas": len(tiendas),
        "totalColaboradores": len(colaboradores),
        "totalOpportunities": len(opportunities),
        "wonOpportunities": len(all_won),
        "activeOpportunities": len(all_active),
        "totalValue": sum(o.get("value", 0) for o in all_won),
        "pipelineValue": sum(o.get("value", 0) for o in all_active),
        "totalProjects": len(projects),
        "totalContacts": len(contacts),
        "conversionRate": round(len(all_won) / len(opportunities) * 100, 1) if opportunities else 0
    }
    
    # Top performers (ordenados por valor total)
    top_performers = sorted(
        metrics_by_user.values(), 
        key=lambda x: x["totalValue"], 
        reverse=True
    )[:10]
    
    return {
        "global": global_metrics,
        "byUser": list(metrics_by_user.values()),
        "topPerformers": top_performers,
        "roleBreakdown": {
            "directores": len(directores),
            "responsables": len(responsables),
            "comerciales": len(comerciales),
            "tiendas": len(tiendas),
            "colaboradores": len(colaboradores)
        }
    }


@api_router.get("/admin/metrics/trends")
async def get_admin_metrics_trends(current_user: dict = Depends(require_admin)):
    """
    [DIRECTOR COMERCIAL ONLY] Métricas de tendencias mensuales.
    Devuelve datos agrupados por mes para gráficos de ventas y pipeline.
    """
    from datetime import timedelta
    from collections import defaultdict
    
    # Get all opportunities
    opportunities = await db.opportunities.find({}, {"_id": 0}).to_list(5000)
    
    # Agrupar por mes
    monthly_data = defaultdict(lambda: {
        "won": 0,
        "wonValue": 0,
        "lost": 0,
        "lostValue": 0,
        "created": 0,
        "createdValue": 0,
        "cocina": 0,
        "cocinaValue": 0,
        "armarios": 0,
        "armariosValue": 0
    })
    
    for opp in opportunities:
        # Extraer mes de creación
        created_at = opp.get("createdAt", "")
        if isinstance(created_at, str) and len(created_at) >= 7:
            month_key = created_at[:7]  # YYYY-MM
        else:
            continue
        
        value = opp.get("value", 0)
        stage = opp.get("stage", "")
        business_type = opp.get("businessType", "cocina")
        
        monthly_data[month_key]["created"] += 1
        monthly_data[month_key]["createdValue"] += value
        
        if stage == "won":
            monthly_data[month_key]["won"] += 1
            monthly_data[month_key]["wonValue"] += value
        elif stage == "lost":
            monthly_data[month_key]["lost"] += 1
            monthly_data[month_key]["lostValue"] += value
        
        # Por tipo de negocio
        if business_type == "cocina":
            monthly_data[month_key]["cocina"] += 1
            monthly_data[month_key]["cocinaValue"] += value
        elif business_type == "armarios":
            monthly_data[month_key]["armarios"] += 1
            monthly_data[month_key]["armariosValue"] += value
    
    # Convertir a lista ordenada
    trends = []
    for month, data in sorted(monthly_data.items()):
        trends.append({
            "month": month,
            "monthLabel": get_month_label(month),
            **data
        })
    
    # Tomar los últimos 12 meses
    trends = trends[-12:] if len(trends) > 12 else trends
    
    # Pipeline por etapa (embudo)
    stage_counts = defaultdict(lambda: {"count": 0, "value": 0})
    for opp in opportunities:
        stage = opp.get("stage", "lead")
        if stage not in ["won", "lost"]:
            stage_counts[stage]["count"] += 1
            stage_counts[stage]["value"] += opp.get("value", 0)
    
    funnel_data = [
        {"stage": "lead", "name": "Nuevo", "count": stage_counts["lead"]["count"], "value": stage_counts["lead"]["value"]},
        {"stage": "contacted", "name": "Contactado", "count": stage_counts["contacted"]["count"], "value": stage_counts["contacted"]["value"]},
        {"stage": "proposal", "name": "Propuesta", "count": stage_counts["proposal"]["count"], "value": stage_counts["proposal"]["value"]},
        {"stage": "negotiation", "name": "Negociación", "count": stage_counts["negotiation"]["count"], "value": stage_counts["negotiation"]["value"]}
    ]
    
    # Distribución por tipo de negocio
    business_type_data = {
        "cocina": {"count": 0, "value": 0, "won": 0, "wonValue": 0},
        "armarios": {"count": 0, "value": 0, "won": 0, "wonValue": 0}
    }
    for opp in opportunities:
        bt = opp.get("businessType", "cocina")
        if bt in business_type_data:
            business_type_data[bt]["count"] += 1
            business_type_data[bt]["value"] += opp.get("value", 0)
            if opp.get("stage") == "won":
                business_type_data[bt]["won"] += 1
                business_type_data[bt]["wonValue"] += opp.get("value", 0)
    
    return {
        "monthly": trends,
        "funnel": funnel_data,
        "byBusinessType": business_type_data
    }


def get_month_label(month_str: str) -> str:
    """Convierte YYYY-MM a etiqueta legible como 'Ene 25'"""
    months_es = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    try:
        year, month = month_str.split("-")
        return f"{months_es[int(month)-1]} {year[2:]}"
    except:
        return month_str


@api_router.get("/commercial/my-shops-work")
async def get_commercial_shops_work(commercial_id: str, current_user: dict = Depends(require_auth)):
    """
    [COMERCIAL ONLY] Obtener todos los trabajos de las tiendas asignadas a este comercial.
    Un usuario sin rol elevado solo puede consultar SUS propias tiendas.
    """
    is_elevated = any(current_user.get(f) for f in (
        "isAdmin", "isResponsableDelegacion", "isGerente", "isDirectorComercial", "isDirectorFabrica"
    ))
    if not is_elevated and commercial_id != current_user.get("id"):
        raise HTTPException(status_code=403, detail="No puedes consultar los trabajos de otro comercial")

    # Get users (shops) assigned to this commercial
    shops = await db.users.find(
        {"linkedRepresentativeId": commercial_id},
        {"_id": 0, "password": 0}
    ).to_list(500)
    
    shop_ids = [s["id"] for s in shops]
    
    # Get projects from these shops
    projects = await db.projects.find(
        {"userId": {"$in": shop_ids}},
        {"_id": 0}
    ).to_list(5000)
    
    # Get opportunities assigned to or created by these shops
    opportunities = await db.opportunities.find(
        {"$or": [
            {"assignedTo": {"$in": shop_ids}},
            {"createdBy": {"$in": shop_ids}}
        ]},
        {"_id": 0}
    ).to_list(5000)
    
    # Enrich with shop info
    shops_map = {s["id"]: s for s in shops}
    
    for proj in projects:
        shop = shops_map.get(proj.get("userId", ""))
        proj["shopName"] = shop.get("clientName", "Desconocido") if shop else "Desconocido"
        proj["shopUsername"] = shop.get("username", "") if shop else ""
    
    for opp in opportunities:
        shop = shops_map.get(opp.get("assignedTo", ""))
        opp["shopName"] = shop.get("clientName", "Sin asignar") if shop else "Sin asignar"
    
    return {
        "shops": shops,
        "projects": projects,
        "opportunities": opportunities,
        "summary": {
            "totalShops": len(shops),
            "totalProjects": len(projects),
            "totalOpportunities": len(opportunities)
        }
    }

@api_router.post("/init")
async def init_data():
    """Comprobar el estado de inicialización de datos base.

    El usuario administrador (master) se crea/actualiza automáticamente en el evento
    de arranque (`startup`) usando MASTER_PASSWORD. Este endpoint ya NO crea usuarios
    con contraseñas débiles; solo informa de si existe un administrador y migra a
    bcrypt cualquier contraseña heredada en texto plano del admin master.
    """
    admin = await db.users.find_one(
        {"$or": [{"id": "user-master-mario"}, {"isAdmin": True}]},
        {"_id": 0}
    )

    if not admin:
        return {
            "message": "No hay administrador. Se creará automáticamente en el próximo "
                       "arranque del servidor (define MASTER_PASSWORD).",
            "initialized": False
        }

    # Migrar contraseña heredada en texto plano del admin a bcrypt, si procede.
    pwd = admin.get("password") or ""
    if pwd and not pwd.startswith("$2"):
        hashed = hash_password(pwd)
        await db.users.update_one({"id": admin["id"]}, {"$set": {"password": hashed}})
        return {"message": "Contraseña del administrador migrada a hash bcrypt", "initialized": True}

    return {"message": "Sistema ya inicializado", "initialized": True}


# ============================================
# DESPIECE (BILL OF MATERIALS) API
# ============================================


def calculate_furniture_despiece(
    item: DespieceItemInput,
    carcass_material: str,
    back_material: str,
    grosor: float,
    back_thickness: float = 8,  # Grosor de trasera en mm
    door_tolerance_height: float = 2,  # Tolerancia alto puerta en mm
    door_tolerance_width: float = 3    # Tolerancia ancho puerta en mm
) -> FurnitureDespiece:
    """
    Calculate the despiece (bill of materials) for a single furniture piece.
    
    TODAS LAS MEDIDAS EN CENTÍMETROS (cm)
    
    REGLA FUNDAMENTAL:
    - VERTICALES (laterales/costados): Usan el ALTO COMPLETO del mueble
    - HORIZONTALES (tapas, estantes): Se les descuenta el GROSOR del casco (x2)
    
    Standard furniture components:
    - LATERAL IZQUIERDO: height x depth (ALTO COMPLETO)
    - LATERAL DERECHO: height x depth (ALTO COMPLETO)
    - TAPA SUPERIOR: (width - 2*grosor) x depth (entre laterales)
    - TAPA INFERIOR: (width - 2*grosor) x depth (entre laterales)
    - TRASERA: (width - 2*grosor) x (height - 0.6cm) (encajada en ranuras)
    - BALDA/ESTANTE: (width - 2*grosor) x (depth - 2cm) - optional shelves
    """
    
    # Todas las dimensiones en cm
    w = float(item.width)   # Ancho en cm
    h = float(item.height)  # Alto en cm  
    d = float(item.depth)   # Fondo en cm
    g = grosor / 10  # Grosor viene en mm, convertir a cm (18mm = 1.8cm)
    back_g = back_thickness / 10  # Grosor trasera en cm (8mm = 0.8cm)
    
    # =============================================
    # REGLAS DE CÁLCULO SEGÚN DOCUMENTO:
    # =============================================
    # 1. LATERALES (verticales): 
    #    - Largo = Alto exterior completo
    #    - Ancho = Fondo exterior - Grosor trasera
    # 
    # 2. HORIZONTALES (tapas, estantes):
    #    - Largo = Ancho exterior - (2 × Grosor lateral)
    #    - Ancho = Fondo exterior - Grosor trasera
    #
    # 3. TRASERA:
    #    - Largo = Ancho exterior - (2 × Grosor lateral) [encaja entre laterales]
    #    - Alto = Alto exterior - margen ranuras (0.6cm)
    #
    # 4. ESTANTES:
    #    - Largo = Ancho interior - margen soportes
    #    - Ancho = Fondo interior - margen frontal
    # =============================================
    
    # Medidas calculadas
    fondo_interior = d - back_g  # Fondo menos trasera para laterales y horizontales
    ancho_interior = w - (2 * g)  # Ancho menos los dos laterales
    
    components = []
    component_id = 0
    
    # Determine furniture type - Soportar nomenclaturas ZC y MV
    code_upper = item.productCode.upper()
    name_upper = item.productName.upper()
    category_upper = item.category.upper()
    
    # ALTOS - ZC: A, 9A | MV: A, ASCE, ASC, AR, ARI, ARU, ARC, AD, AV, AE, AM, AMF, ACA, ACC, ASF, AT, ATP, AA, AC, ACP, ACPJ
    #         MV también: L (ALTILLO), LV, S (SOBREENCIMERA), SV, SC, SVC, BOA, BOS
    alto_prefixes = ['ASCE', 'ASC', 'ARI', 'ARU', 'ARC', 'AD', 'AV', 'AE', 'AMF', 'AM', 'ACA', 'ACC', 'ASF', 'ATP', 'AT', 'AA', 'ACPJ', 'ACP', 'AC', 'AR', 'LD', 'LV', 'SVC', 'SV', 'SC', 'BOA', 'BOS']
    is_alto = (
        "ALTO" in name_upper or "ALTO" in category_upper or
        any(code_upper.startswith(p) for p in alto_prefixes) or
        code_upper.startswith('9A') or
        (code_upper.startswith('A') and len(code_upper) > 1 and code_upper[1].isdigit()) or
        (code_upper.startswith('L') and len(code_upper) > 1 and code_upper[1].isdigit()) or
        (code_upper.startswith('S') and len(code_upper) > 1 and code_upper[1].isdigit()) or
        "ALTILLO" in name_upper or "SOBREENCIMERA" in name_upper
    )
    
    # BAJOS - ZC: B, 9B | MV: B, BF, BRI, BRU, BR, BH, BHC, BHZ, BHG, BT, BTP, BPC, BC, BCG, BGC, BCGF, BGF
    bajo_prefixes = ['BRI', 'BRU', 'BR', 'BHZ', 'BHG', 'BHC', 'BH', 'BTP', 'BT', 'BPC', 'BCGF', 'BCG', 'BGF', 'BGC', 'BC', 'BF']
    is_bajo = (
        "BAJO" in name_upper or "BAJO" in category_upper or
        any(code_upper.startswith(p) for p in bajo_prefixes) or
        code_upper.startswith('9B') or
        (code_upper.startswith('B') and len(code_upper) > 1 and code_upper[1].isdigit()) or
        "FREGADERO" in name_upper or "HORNO" in name_upper
    )
    
    # COLUMNAS - ZC: C | MV: CD, CE, CF, CH, CHPC, CHGC, CHC, CHM, CHMG, CHMC, CHMCG, BOC
    #            MV también: M (MEDIA COLUMNA), MV, MPG, MVG, MPH, MPM, MGHM, MCHM
    columna_prefixes = ['CD', 'CE', 'CF', 'CHPC', 'CHGC', 'CHC', 'CHMCG', 'CHMG', 'CHMC', 'CHM', 'CH', 'BOC', 'MGHM', 'MCHM', 'MPG', 'MVG', 'MPH', 'MPM']
    is_columna = (
        "COLUMNA" in name_upper or "COLUMNA" in category_upper or
        "SEMICOLUMNA" in name_upper or "MEDIACOLUMNA" in name_upper or
        any(code_upper.startswith(p) for p in columna_prefixes) or
        (code_upper.startswith('C') and len(code_upper) > 1 and code_upper[1].isdigit()) or
        (code_upper.startswith('M') and len(code_upper) > 1 and code_upper[1].isdigit()) or
        (code_upper.startswith('MV') and len(code_upper) > 2 and code_upper[2].isdigit())
    )
    
    # Tipos especiales
    is_fregadero = "FREGADERO" in name_upper or "FREG" in code_upper
    is_horno = "HORNO" in name_upper or "BH" in code_upper[:2]
    is_rincon = "RINCON" in name_upper or "RINCÓN" in name_upper or code_upper.startswith('BR') or code_upper.startswith('AR')
    
    def add_component(name: str, short: str, material: str, length_cm: float, width_cm: float, thickness_cm: float = None, qty: int = 1, notes: str = ""):
        nonlocal component_id
        component_id += 1
        if thickness_cm is None:
            thickness_cm = g
        # Área en m² (cm² / 10000)
        area = (length_cm * width_cm * qty) / 10_000
        components.append(ComponentPiece(
            id=f"CMP-{item.productId[:8]}-{component_id:03d}",
            name=name,
            nameShort=short,
            material=material,
            length=round(length_cm, 1),  # cm
            width=round(width_cm, 1),    # cm
            thickness=round(thickness_cm * 10, 1),  # Mostrar grosor en mm
            quantity=qty,
            area=round(area, 4),
            notes=notes
        ))
    
    # =============================================
    # VERTICALES - LATERALES (costados)
    # Largo = ALTO COMPLETO del mueble
    # Ancho = FONDO - GROSOR TRASERA
    # =============================================
    
    # LATERAL IZQUIERDO (Side panel - full height)
    # Para muebles de rincón, el lateral tiene forma especial
    lateral_fondo = fondo_interior if not is_rincon else fondo_interior - 5  # Rincón: lateral más corto
    add_component(
        "Lateral izquierdo", "LAT-I",
        carcass_material,
        h, lateral_fondo, g, 1,  # ALTO COMPLETO x (FONDO - TRASERA)
        "1L canto" + (" (rincón)" if is_rincon else "")
    )
    
    # LATERAL DERECHO (Side panel - full height)
    add_component(
        "Lateral derecho", "LAT-D",
        carcass_material,
        h, lateral_fondo, g, 1,  # ALTO COMPLETO x (FONDO - TRASERA)
        "1L canto" + (" (rincón)" if is_rincon else "")
    )
    
    # =============================================
    # HORIZONTALES - TAPAS (superior e inferior)
    # Largo = ANCHO - (2 × GROSOR LATERAL) = ancho_interior
    # Ancho = FONDO - GROSOR TRASERA = fondo_interior
    # =============================================
    
    # TAPA SUPERIOR (Top panel - between sides)
    # Los fregaderos NO tienen tapa superior (hueco para la encimera)
    if not is_fregadero:
        add_component(
            "Tapa superior", "TAPA-S", 
            carcass_material,
            ancho_interior, fondo_interior, g, 1,
            "1L canto"
        )
    else:
        # Fregadero: travesaños frontales y traseros en lugar de tapa
        add_component(
            "Travesaño frontal", "TRAV-F",
            carcass_material,
            ancho_interior, 8, g, 1,  # 8cm de ancho
            "Travesaño para fregadero"
        )
        add_component(
            "Travesaño trasero", "TRAV-T",
            carcass_material,
            ancho_interior, 8, g, 1,
            "Travesaño para fregadero"
        )
    
    # TAPA INFERIOR (Bottom panel) - varía según tipo de mueble
    if is_alto:
        # ALTOS usan un travesaño inferior estrecho (8cm de ancho)
        add_component(
            "Travesaño inferior", "TRAV-I",
            carcass_material,
            ancho_interior, 8, g, 1,
            ""
        )
    elif is_horno:
        # Hornos: travesaños laterales para soporte del horno
        add_component(
            "Travesaño inferior horno", "TRAV-H",
            carcass_material,
            ancho_interior, 10, g, 1,
            "Soporte para horno empotrado"
        )
    else:
        # Bajos y columnas: tapa inferior completa
        add_component(
            "Tapa inferior", "TAPA-I",
            carcass_material,
            ancho_interior, fondo_interior, g, 1,
            "1L canto" if is_bajo else ""
        )
    
    # =============================================
    # TRASERA (Back panel)
    # Largo = ANCHO INTERIOR (entre laterales)
    # Alto = ALTO - margen ranuras (0.6cm total)
    # =============================================
    back_height = h - 0.6  # Encajada en ranuras (0.3cm arriba y abajo)
    add_component(
        "Trasera modulo", "TRAS",
        back_material,
        ancho_interior, back_height, back_g, 1,
        f"Tablero {int(back_thickness)}mm"
    )
    
    # =============================================
    # CAJONES - Detección (debe calcularse ANTES de las baldas, que dependen de has_drawers)
    # =============================================
    has_drawers = False
    num_drawers = 0
    drawer_height_cm = 0

    # Detectar cajones por código/nombre
    cajon_keywords = ['CAJ', 'CAJON', 'CAJÓN', 'DRAWER', 'BT', 'BCG', 'BCGF', 'BGF', 'BGC']
    if any(k in code_upper or k in name_upper for k in cajon_keywords):
        has_drawers = True
        # Estimar número de cajones según la altura
        if h >= 70:
            num_drawers = 3
        elif h >= 50:
            num_drawers = 2
        else:
            num_drawers = 1
        drawer_height_cm = round((h - 1.0) / num_drawers - 0.3, 1)  # Alto interior cajón

    # =============================================
    # BALDAS / ESTANTES (Shelves)
    # Largo = ANCHO INTERIOR - margen soportes (0.5cm)
    # Ancho = FONDO INTERIOR - margen frontal (2cm retranqueo)
    # =============================================
    shelf_count = 0
    
    # Columnas: más estantes debido a su altura
    if is_columna:
        if h >= 200:  # Columna completa
            shelf_count = 5
        elif h >= 140:  # Media columna
            shelf_count = 3
        else:
            shelf_count = 2
    # Altos: normalmente 1-2 estantes
    elif is_alto:
        if h >= 90:
            shelf_count = 2
        elif h >= 60:
            shelf_count = 1
        else:
            shelf_count = 0
    # Bajos: normalmente 1 estante
    elif is_bajo:
        if is_fregadero or is_horno or has_drawers:
            shelf_count = 0  # Fregaderos, hornos y cajones sin estantes
        elif h >= 80:
            shelf_count = 2  # Bajos altos (80cm+) llevan 2 estantes
        else:
            shelf_count = 1  # Bajos estándar (70-72cm) solo 1 estante
    # Genérico: basado en altura
    elif h >= 70:
        shelf_count = 2 if h < 120 else 3 if h < 180 else 4
    elif h >= 35:
        shelf_count = 1
    
    if shelf_count > 0:
        shelf_length = ancho_interior - 0.5  # Entre laterales menos margen soportes
        shelf_width = fondo_interior - 2  # Retranqueado 2cm del frontal
        add_component(
            "Balda interior", "BALDA",
            carcass_material,
            shelf_length, shelf_width, g, shelf_count,
            "Regulable con soportes"
        )
    
    # =============================================
    # PUERTAS - Calcular dimensiones de puertas
    # =============================================
    # Detectar si el mueble tiene puertas según su nombre/categoría/código
    
    # Detectar cantidad de puertas
    has_doors = False
    num_doors = 0
    
    # Patrones para detectar puertas:
    # ZC: 1P, 2P, 3P, 4P (ej: 9A1P300, 9B2P600)
    # MV: Similar pero también D/I (derecha/izquierda)
    
    # 4P = 4 puertas (columnas grandes)
    if "4P" in name_upper or "4P" in code_upper:
        has_doors = True
        num_doors = 4
    # 3P = 3 puertas
    elif "3P" in name_upper or "3P" in code_upper:
        has_doors = True
        num_doors = 3
    # 2P = 2 puertas
    elif "2P" in name_upper or "2P" in code_upper:
        has_doors = True
        num_doors = 2
    # 1P = 1 puerta
    elif "1P" in name_upper or "1P" in code_upper:
        has_doors = True
        num_doors = 1
    # D/I = 1 puerta (derecha/izquierda)
    elif "/I" in name_upper or "/D" in name_upper or "D/I" in name_upper:
        has_doors = True
        num_doors = 1
    elif "PUERTA" in name_upper or "PTA" in code_upper:
        has_doors = True
        num_doors = 1
    # Muebles especiales SIN puertas
    elif is_fregadero or "MICROONDAS" in name_upper or "ESTANTERIA" in name_upper or "HUECO" in name_upper:
        has_doors = False
        num_doors = 0
    # Muebles que normalmente tienen puertas pero no lo especifican
    elif is_alto or is_bajo or is_columna:
        has_doors = True
        # Estimar número de puertas según el ancho
        if w <= 45:
            num_doors = 1
        elif w <= 90:
            num_doors = 2
        elif w <= 120:
            num_doors = 3
        else:
            num_doors = 4
    
    # =============================================
    # CAJONES - Alta de componentes (la detección se hizo antes de las baldas)
    # =============================================
    if has_drawers and num_drawers > 0:
        drawer_width = ancho_interior - 2  # Ancho cajón = ancho interior menos guías
        drawer_depth = fondo_interior - 4  # Fondo cajón = fondo interior menos frontal

        # Frentes de cajón
        add_component(
            f"Frente cajón {'(' + str(num_drawers) + ' uds)' if num_drawers > 1 else ''}",
            "FRENTE-CAJ",
            "PUERTA COLOR",
            round(drawer_height_cm, 1), round(w - 0.3, 1), 1.9, num_drawers,
            f"Frente cajón acabado a elegir. {num_drawers} ud{'s' if num_drawers > 1 else ''}"
        )
        # Laterales cajón (x2 por cajón)
        add_component(
            "Laterales cajón",
            "LAT-CAJ",
            "Tablero 16mm",
            round(drawer_depth, 1), round(drawer_height_cm - 1, 1), 1.6, num_drawers * 2,
            "2 laterales por cajón"
        )
        # Fondo cajón
        add_component(
            "Fondo cajón",
            "FOND-CAJ",
            "Tablero 8mm",
            round(drawer_width - 1, 1), round(drawer_depth, 1), 0.8, num_drawers,
            "Fondo 8mm por cajón"
        )
        # Guías cajón
        add_component(
            "Guías cajón",
            "GUIAS",
            "HERRAJE",
            0, 0, 0, num_drawers,
            f"1 juego de guías por cajón ({num_drawers} juegos)"
        )

    # Calcular dimensiones de puerta si tiene puertas
    if has_doors and num_doors > 0:
        # =============================================
        # PUERTAS - Reglas según tipo de mueble:
        #
        # Columnas 4P: 2 puertas arriba + 2 puertas abajo
        #   → door_height = h/2 - tolerancia (NO h completo)
        # Columnas 2P: puerta única a toda la altura
        # Altos/Bajos: altura completa del mueble
        # =============================================
        door_height_tolerance = door_tolerance_height / 10  # mm → cm
        door_width_tolerance = door_tolerance_width / 10    # mm → cm
        door_gap_between = door_width_tolerance
        door_thickness_cm = 1.9

        # ── Alto de puerta ──────────────────────────
        if is_columna and num_doors == 4:
            # 4 puertas en 2 filas: cada puerta cubre la mitad de la altura
            door_height = round(h / 2 - door_height_tolerance, 1)
            door_rows = 2   # 2 filas de puertas
        elif is_columna and num_doors == 6:
            # 6 puertas en 3 filas (columnas muy altas)
            door_height = round(h / 3 - door_height_tolerance, 1)
            door_rows = 3
        else:
            # 1P, 2P, 3P → siempre la altura completa del mueble
            door_height = round(h - door_height_tolerance, 1)
            door_rows = 1

        # ── Ancho de puerta ─────────────────────────
        # En columnas 4P: 2 puertas por fila
        doors_per_row = num_doors // door_rows if door_rows > 1 else num_doors
        if doors_per_row == 1:
            door_width = round(w - door_width_tolerance, 1)
        elif doors_per_row == 2:
            door_width = round((w - door_gap_between) / 2, 1)
        elif doors_per_row == 3:
            door_width = round((w - 2 * door_gap_between) / 3, 1)
        else:
            door_width = round((w - (doors_per_row - 1) * door_gap_between) / doors_per_row, 1)

        # Descripción clara
        door_desc = f"{num_doors} puerta{'s' if num_doors > 1 else ''} acabado a elegir"
        if door_rows > 1:
            door_desc += f" ({door_rows} filas × {doors_per_row} puertas, {door_height}cm alto c/u)"

        add_component(
            f"Puerta {'(' + str(num_doors) + ' uds)' if num_doors > 1 else ''}",
            "PUERTA",
            "PUERTA COLOR",
            door_height, door_width, door_thickness_cm, num_doors,
            door_desc
        )
    
    # =============================================
    # HERRAJES - Según tipo de mueble
    # =============================================
    
    # COLGADORES - Solo para muebles ALTOS (1 juego = 2 colgadores)
    if is_alto:
        add_component(
            "Juego de colgadores",
            "COLG",
            "HERRAJE",
            0, 0, 0, 1,
            "1 juego = 2 colgadores para mueble alto de pared"
        )
    
    # BISAGRAS - Según altura REAL de la puerta (no del mueble)
    if has_doors and num_doors > 0:
        # En columnas 4P door_height = h/2, en el resto = h completo
        real_door_h = door_height if 'door_height' in locals() else h
        bisagras_por_puerta = 2 if real_door_h <= 90 else 3
        total_bisagras = bisagras_por_puerta * num_doors
        add_component(
            "Bisagras",
            "BISAG",
            "HERRAJE",
            0, 0, 0, total_bisagras,
            f"{bisagras_por_puerta} bisagras/puerta (puerta {real_door_h}cm)"
        )
        
        # TIRADORES - 1 por puerta
        add_component(
            "Tiradores",
            "TIRAD",
            "HERRAJE",
            0, 0, 0, num_doors,
            "1 tirador por puerta"
        )
    
    # SOPORTES DE BALDA - 4 por balda
    if shelf_count > 0:
        add_component(
            "Soportes de balda",
            "SOPORT",
            "HERRAJE",
            0, 0, 0, shelf_count * 4,
            "4 soportes por balda"
        )
    
    # PATAS/ZÓCALO - Solo para BAJOS y COLUMNAS
    if is_bajo or is_columna:
        # 4 patas para muebles estándar, 6 para anchos > 80cm
        num_patas = 6 if w > 80 else 4
        add_component(
            "Patas regulables",
            "PATAS",
            "HERRAJE",
            0, 0, 0, num_patas,
            f"{num_patas} patas de 10-15cm regulables"
        )
    
    # Calculate totals
    total_panels = sum(c.quantity for c in components)
    total_area = sum(c.area for c in components)
    
    return FurnitureDespiece(
        productId=item.productId,
        productCode=item.productCode,
        productName=item.productName,
        category=item.category,
        originalWidth=item.width,
        originalHeight=item.height,
        originalDepth=item.depth,
        itemQuantity=item.quantity,
        components=components,
        totalPanels=total_panels,
        totalArea=round(total_area * item.quantity, 4)
    )

@api_router.post("/despiece/calculate", response_model=DespieceResponse)
async def calculate_despiece(request: DespieceRequest):
    """
    Calculate the bill of materials (despiece) for a list of furniture items.
    Returns cutting lists and assembly orders.
    """
    try:
        despiece_items = []
        
        for item in request.items:
            furniture_despiece = calculate_furniture_despiece(
                item,
                request.carcassMaterial,
                request.backPanelMaterial,
                request.grosor,
                request.backThickness,
                request.doorToleranceHeight,
                request.doorToleranceWidth
            )
            despiece_items.append(furniture_despiece)
        
        # Calculate summary
        total_pieces = sum(d.totalPanels * d.itemQuantity for d in despiece_items)
        total_area = sum(d.totalArea for d in despiece_items)
        total_furniture = sum(d.itemQuantity for d in despiece_items)
        
        # Group by material
        material_summary = {}
        for d in despiece_items:
            for comp in d.components:
                mat = comp.material
                if mat not in material_summary:
                    material_summary[mat] = {"pieces": 0, "area": 0}
                material_summary[mat]["pieces"] += comp.quantity * d.itemQuantity
                material_summary[mat]["area"] += comp.area * d.itemQuantity
        
        # Round areas
        for mat in material_summary:
            material_summary[mat]["area"] = round(material_summary[mat]["area"], 3)
        
        return DespieceResponse(
            items=despiece_items,
            summary={
                "totalFurniture": total_furniture,
                "totalPieces": total_pieces,
                "totalArea": round(total_area, 3),
                "byMaterial": material_summary
            },
            generatedAt=datetime.now(timezone.utc).isoformat()
        )
    except Exception as e:
        logger.error(f"Calculate despiece error: {e}")
        raise HTTPException(status_code=500, detail=f"Error calculando despiece: {str(e)}")
# ============================================
# DATABASE EXPORT API (Admin Only)
# ============================================

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io

@api_router.get("/admin/export-database")
async def export_database(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """
    Exportar toda la base de datos a Excel.
    Solo accesible para Director Comercial (admin).
    """
    if not credentials:
        raise HTTPException(status_code=401, detail="Token requerido")
    
    try:
        # Verify admin role
        payload = verify_access_token(credentials.credentials)
        user_id = payload.get("sub")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        
        if not user or not (user.get('isAdmin') or user.get('isDirectorComercial')):
            raise HTTPException(status_code=403, detail="Solo el Director Comercial o un administrador puede exportar la base de datos")
        
        # Create workbook
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        # ===== USUARIOS =====
        ws_users = wb.active
        ws_users.title = "Usuarios"
        
        users = await db.users.find({}, {'_id': 0, 'password': 0}).to_list(1000)
        user_headers = ['ID', 'Usuario', 'Nombre', 'Rol', 'Email', 'Teléfono', 'Delegación', 'Activo']
        for col, header in enumerate(user_headers, 1):
            cell = ws_users.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_num, u in enumerate(users, 2):
            ws_users.cell(row=row_num, column=1, value=u.get('id', ''))
            ws_users.cell(row=row_num, column=2, value=u.get('username', ''))
            ws_users.cell(row=row_num, column=3, value=u.get('fullName', ''))
            ws_users.cell(row=row_num, column=4, value=u.get('role', ''))
            ws_users.cell(row=row_num, column=5, value=u.get('email', ''))
            ws_users.cell(row=row_num, column=6, value=u.get('phone', ''))
            ws_users.cell(row=row_num, column=7, value=u.get('delegation', ''))
            ws_users.cell(row=row_num, column=8, value='Sí' if u.get('isActive', True) else 'No')
        
        # ===== PRODUCTOS MONTADA =====
        ws_products = wb.create_sheet("Productos Montada")
        
        products = await db.products.find({}, {'_id': 0}).to_list(10000)
        prod_headers = ['Código', 'Nombre', 'Programa', 'Categoría', 'Serie', 'Ancho', 'Alto', 'Fondo', 'Puntos', 'Z1', 'Z2', 'Z3', 'Z4']
        for col, header in enumerate(prod_headers, 1):
            cell = ws_products.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_num, p in enumerate(products, 2):
            zp = p.get('zonePoints', {}) or {}
            ws_products.cell(row=row_num, column=1, value=p.get('code', ''))
            ws_products.cell(row=row_num, column=2, value=p.get('name', ''))
            ws_products.cell(row=row_num, column=3, value=p.get('programa', ''))
            ws_products.cell(row=row_num, column=4, value=p.get('category', ''))
            ws_products.cell(row=row_num, column=5, value=p.get('series', ''))
            ws_products.cell(row=row_num, column=6, value=p.get('width', 0))
            ws_products.cell(row=row_num, column=7, value=p.get('height', 0))
            ws_products.cell(row=row_num, column=8, value=p.get('depth', 0))
            ws_products.cell(row=row_num, column=9, value=p.get('points', 0))
            ws_products.cell(row=row_num, column=10, value=zp.get('Z1', 0))
            ws_products.cell(row=row_num, column=11, value=zp.get('Z2', 0))
            ws_products.cell(row=row_num, column=12, value=zp.get('Z3', 0))
            ws_products.cell(row=row_num, column=13, value=zp.get('Z4', 0))
        
        # ===== PRODUCTOS DESPIECE (Tableros) =====
        ws_despiece = wb.create_sheet("Productos Despiece")
        
        despiece_products = await db.despiece_products.find({}, {'_id': 0}).to_list(10000)
        desp_headers = ['Código', 'Nombre', 'Fabricante', 'Colección', 'Color', 'Acabado', 'Grosor', 'Categoría', 'Precio Z1', 'Precio Z2', 'Precio Z3']
        for col, header in enumerate(desp_headers, 1):
            cell = ws_despiece.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_num, p in enumerate(despiece_products, 2):
            ws_despiece.cell(row=row_num, column=1, value=p.get('code', ''))
            ws_despiece.cell(row=row_num, column=2, value=p.get('name', ''))
            ws_despiece.cell(row=row_num, column=3, value=p.get('manufacturer', ''))
            ws_despiece.cell(row=row_num, column=4, value=p.get('collection', ''))
            ws_despiece.cell(row=row_num, column=5, value=p.get('color', ''))
            ws_despiece.cell(row=row_num, column=6, value=p.get('finish', ''))
            ws_despiece.cell(row=row_num, column=7, value=p.get('thickness', 0))
            ws_despiece.cell(row=row_num, column=8, value=p.get('category', ''))
            ws_despiece.cell(row=row_num, column=9, value=p.get('priceZ1', 0))
            ws_despiece.cell(row=row_num, column=10, value=p.get('priceZ2', 0))
            ws_despiece.cell(row=row_num, column=11, value=p.get('priceZ3', 0))
        
        # ===== PROYECTOS =====
        ws_projects = wb.create_sheet("Proyectos")
        
        projects = await db.projects.find({}, {'_id': 0, 'itemsMontada': 0, 'itemsDespiece': 0}).to_list(1000)
        proj_headers = ['Nº Expediente', 'Cliente', 'Dirección', 'Total PVP', 'Estado', 'Comercial']
        for col, header in enumerate(proj_headers, 1):
            cell = ws_projects.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_num, p in enumerate(projects, 2):
            ws_projects.cell(row=row_num, column=1, value=p.get('budgetNumber', ''))
            ws_projects.cell(row=row_num, column=2, value=p.get('customerName', ''))
            ws_projects.cell(row=row_num, column=3, value=p.get('customerAddress', ''))
            ws_projects.cell(row=row_num, column=4, value=p.get('totalPvp', 0))
            ws_projects.cell(row=row_num, column=5, value=p.get('status', 'activo'))
            ws_projects.cell(row=row_num, column=6, value=p.get('savedBy', ''))
        
        # Adjust column widths
        for ws in wb.worksheets:
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Log audit
        audit.log(
            AuditAction.PROJECT_EXPORT,
            user_id=user_id,
            username=user.get("username"),
            resource_type="database",
            success=True,
            details={"tables": ["users", "products_montada", "products_despiece", "projects"]}
        )
        
        filename = f"LUIGGI_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Database export error: {e}")
        raise HTTPException(status_code=500, detail=f"Error exportando base de datos: {str(e)}")

@api_router.post("/products/fix-data")
async def fix_product_data(current_user: dict = Depends(require_admin)):
    """
    Fix product data:
    1. Semicolumnas series: 11 -> 110, 12 -> 120, etc.
    2. Bajos fondo: 33 -> 58 (para muebles bajos estándar)
    """
    try:
        results = {
            "series_fixed": 0,
            "depth_fixed": 0,
            "errors": []
        }
        
        # 1. Fix semicolumnas series names
        series_mapping = {
            "SEMICOLUMNAS 11 FONDO ESTANDAR": "SEMICOLUMNAS 110 FONDO ESTANDAR",
            "SEMICOLUMNAS 12 FONDO ESTANDAR": "SEMICOLUMNAS 120 FONDO ESTANDAR",
            "SEMICOLUMNAS 13 FONDO ESTANDAR": "SEMICOLUMNAS 130 FONDO ESTANDAR",
            "SEMICOLUMNAS 14 FONDO ESTANDAR": "SEMICOLUMNAS 140 FONDO ESTANDAR",
            "SEMICOLUMNAS 16 FONDO ESTANDAR": "SEMICOLUMNAS 160 FONDO ESTANDAR",
            "GOLA - SEMICOLUMNAS 11 FONDO ESTANDAR": "GOLA - SEMICOLUMNAS 110 FONDO ESTANDAR",
            "GOLA - SEMICOLUMNAS 12 FONDO ESTANDAR": "GOLA - SEMICOLUMNAS 120 FONDO ESTANDAR",
            "GOLA - SEMICOLUMNAS 13 FONDO ESTANDAR": "GOLA - SEMICOLUMNAS 130 FONDO ESTANDAR",
            "GOLA - SEMICOLUMNAS 14 FONDO ESTANDAR": "GOLA - SEMICOLUMNAS 140 FONDO ESTANDAR",
            "GOLA - SEMICOLUMNAS 16 FONDO ESTANDAR": "GOLA - SEMICOLUMNAS 160 FONDO ESTANDAR",
        }
        
        for old_series, new_series in series_mapping.items():
            result = await db.products.update_many(
                {"series": old_series},
                {"$set": {"series": new_series}}
            )
            results["series_fixed"] += result.modified_count
            logger.info(f"Fixed series: {old_series} -> {new_series} ({result.modified_count} products)")
        
        # 2. Fix bajos fondo - Muebles BAJOS con fondo 33 -> 58
        # Pero NO los ALTOS que tienen fondo 33 correctamente
        # Criteria: codigo empieza con 7B, 8B (bajos 70cm, 80cm) y NO son ALTOS
        bajo_codes = ["7B", "8B"]  # Códigos que empiezan con estos son BAJOS
        
        for code_prefix in bajo_codes:
            result = await db.products.update_many(
                {
                    "code": {"$regex": f"^{code_prefix}"},
                    "depth": 33.0
                },
                {"$set": {"depth": 58.0}}
            )
            results["depth_fixed"] += result.modified_count
            logger.info(f"Fixed depth for {code_prefix}* products: {result.modified_count}")
        
        # También buscar productos con "Bajo" en el nombre que tengan fondo 33
        result = await db.products.update_many(
            {
                "name": {"$regex": "^Bajo", "$options": "i"},
                "depth": 33.0
            },
            {"$set": {"depth": 58.0}}
        )
        results["depth_fixed"] += result.modified_count
        
        return {
            "success": True,
            "message": "Datos de productos corregidos",
            "results": results
        }
    except Exception as e:
        logger.error(f"Error fixing product data: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Global Rate Limiting Middleware - protege TODOS los endpoints automáticamente
app.add_middleware(GlobalRateLimitMiddleware, default_rpm=60)

# CORS Middleware - MUST be added BEFORE routers
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', 'https://erp.luiggihome.es,https://erp-crm-14-03-2026-production.up.railway.app,http://localhost:3000').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Compresión Gzip: respuestas mas ligeras (catalogos, presupuestos, listados
# grandes cargan mucho mas rapido). Solo comprime respuestas > 500 bytes.
app.add_middleware(GZipMiddleware, minimum_size=500)

# Include the router in the main app (AFTER middleware)
app.include_router(api_router)

# ============================================
# STARTUP / SHUTDOWN EVENTS
# ============================================

@app.on_event("startup")
async def startup_event():
    """Configure scheduled backups and database indexes on startup"""
    # =============================================
    # ÍNDICES DE BASE DE DATOS - Fortalecer integridad
    # =============================================
    try:
        # Índice único para expedientes
        await db.digitalizador_history.create_index("expNumber", unique=True, sparse=True)
        logger.info("Índice único creado para expNumber en digitalizador_history")
        
        # Índices para usuarios
        await db.users.create_index("username", unique=True)
        await db.users.create_index("email", sparse=True)
        await db.users.create_index("factoryId", sparse=True)
        logger.info("Índices de usuarios creados")
        
        # Índices para pedidos (fabrica_orders)
        await db.fabrica_orders.create_index("id", unique=True)
        await db.fabrica_orders.create_index("budgetNumber")
        await db.fabrica_orders.create_index("status")
        await db.fabrica_orders.create_index("factoryId", sparse=True)
        await db.fabrica_orders.create_index("createdAt")
        await db.fabrica_orders.create_index([("status", 1), ("factoryId", 1)])  # Índice compuesto
        logger.info("Índices de fabrica_orders creados")
        
        # Índices para presupuestos
        await db.budgets.create_index("id", unique=True)
        await db.budgets.create_index("createdAt")
        await db.budgets.create_index("userId", sparse=True)
        logger.info("Índices de budgets creados")
        
        # Índices para clientes
        await db.clients.create_index("id", unique=True)
        await db.clients.create_index("name")
        await db.clients.create_index("email", sparse=True)
        logger.info("Índices de clients creados")
        
        # Índices para catálogos y productos
        await db.catalogs.create_index("id", unique=True)
        await db.catalogs.create_index("libraryId")
        logger.info("Índices de catalogs creados")
        
        # Índices para fábricas
        await db.factories.create_index("id", unique=True)
        await db.factories.create_index("code", unique=True)
        logger.info("Índices de factories creados")
        
        # Índices para historial de cambios (order_history)
        await db.order_history.create_index("orderId")
        await db.order_history.create_index("timestamp")
        await db.order_history.create_index([("orderId", 1), ("timestamp", -1)])
        logger.info("Índices de order_history creados")

        # Índices para CRM - contacts
        await db.contacts.create_index("id", unique=True)
        await db.contacts.create_index("createdByUserId", sparse=True)
        await db.contacts.create_index("assignedToId", sparse=True)
        await db.contacts.create_index("status")
        await db.contacts.create_index("name")
        await db.contacts.create_index([("assignedToId", 1), ("status", 1)])
        logger.info("Índices de contacts creados")

        # Índices para CRM - opportunities
        await db.opportunities.create_index("id", unique=True)
        await db.opportunities.create_index("contactId")
        await db.opportunities.create_index("stage")
        await db.opportunities.create_index("assignedTo", sparse=True)
        await db.opportunities.create_index("createdByUserId", sparse=True)
        await db.opportunities.create_index([("stage", 1), ("assignedTo", 1)])
        await db.opportunities.create_index("updatedAt")
        logger.info("Índices de opportunities creados")

        # Índices para projects (presupuestos)
        await db.projects.create_index("id", unique=True)
        await db.projects.create_index("userId")
        await db.projects.create_index("budgetNumber")
        await db.projects.create_index("status")
        await db.projects.create_index("clientCode", sparse=True)
        await db.projects.create_index("validUntil", sparse=True)
        await db.projects.create_index([("userId", 1), ("status", 1)])
        await db.projects.create_index("updatedAt")
        logger.info("Índices de projects creados")

        # Índices para products (catálogo)
        await db.products.create_index("id", unique=True)
        await db.products.create_index("code")
        await db.products.create_index("library")
        await db.products.create_index("category")
        await db.products.create_index([("library", 1), ("category", 1)])
        await db.products.create_index([("library", 1), ("code", 1)])
        logger.info("Índices de products creados")

        # Índices para despiece_products
        await db.despiece_products.create_index("id", unique=True)
        await db.despiece_products.create_index("manufacturer")
        await db.despiece_products.create_index("collection", sparse=True)
        logger.info("Índices de despiece_products creados")

        # Índices para facturas
        await db.invoices.create_index("id", unique=True)
        await db.invoices.create_index("status")
        await db.invoices.create_index("clientName")
        await db.invoices.create_index("invoiceNumber")
        await db.invoices.create_index("projectId", sparse=True)
        await db.invoices.create_index("dueDate", sparse=True)
        await db.invoices.create_index("createdAt")
        logger.info("Índices de invoices creados")

        # Índices para CRM - contactos, oportunidades, actividades
        await db.contacts.create_index("id", unique=True)
        await db.contacts.create_index("createdByUserId", sparse=True)
        await db.contacts.create_index("status")
        await db.contacts.create_index([("createdByUserId", 1), ("status", 1)])
        await db.opportunities.create_index("id", unique=True)
        await db.opportunities.create_index("contactId", sparse=True)
        await db.opportunities.create_index("assignedTo", sparse=True)
        await db.opportunities.create_index("stage")
        await db.opportunities.create_index([("assignedTo", 1), ("stage", 1)])
        await db.activities.create_index("id", unique=True)
        await db.activities.create_index("contactId", sparse=True)
        await db.activities.create_index("opportunityId", sparse=True)
        await db.activities.create_index("userId", sparse=True)
        await db.activities.create_index("createdAt")
        logger.info("Índices de CRM creados")

        # Índices para Command Center - actividad de usuarios
        await db.user_activity.create_index("userId")
        await db.user_activity.create_index("activityType")
        await db.user_activity.create_index("timestamp")
        await db.user_activity.create_index([("userId", 1), ("timestamp", -1)])
        await db.user_activity.create_index([("activityType", 1), ("timestamp", -1)])
        logger.info("Índices de user_activity creados")

        # Índices para projects (presupuestos) - rendimiento panel mando
        await db.projects.create_index("userId", sparse=True)
        await db.projects.create_index("status")
        await db.projects.create_index([("userId", 1), ("status", 1)])
        await db.projects.create_index([("userId", 1), ("createdAt", -1)])
        logger.info("Índices de projects creados")

        # Índices para montajes
        await db.montajes.create_index("id", unique=True)
        await db.montajes.create_index("status")
        await db.montajes.create_index("assignedTo", sparse=True)
        await db.montajes.create_index("scheduledDate", sparse=True)
        await db.montajes.create_index([("status", 1), ("scheduledDate", 1)])
        logger.info("Índices de montajes creados")

        # Índices para audit_log (seguridad y trazabilidad)
        await db.audit_log.create_index("userId", sparse=True)
        await db.audit_log.create_index("action")
        await db.audit_log.create_index("timestamp")
        await db.audit_log.create_index([("userId", 1), ("timestamp", -1)])
        await db.audit_log.create_index([("action", 1), ("timestamp", -1)])
        logger.info("Índices de audit_log creados")

        # Índices para settings
        await db.settings.create_index("key", unique=True)
        logger.info("Índices de settings creados")

        # Índices para libraries
        await db.libraries.create_index("id", unique=True)
        await db.libraries.create_index("code", unique=True)
        logger.info("Índices de libraries creados")

        # Índices para distributor_requests
        await db.distributor_requests.create_index("status")
        await db.distributor_requests.create_index("createdAt")
        logger.info("Índices de distributor_requests creados")

        # Text search indexes para búsqueda rápida
        await db.products.create_index([("name", "text"), ("description", "text"), ("code", "text")])
        await db.clients.create_index([("name", "text"), ("email", "text"), ("phone", "text")])
        await db.contacts.create_index([("name", "text"), ("email", "text"), ("company", "text")])
        logger.info("Índices de texto completo creados")

        # Índices para colecciones de rendimiento adicionales (rentabilidad,
        # calendario, agenda prescriptor y Google Calendar)
        await db.project_costs.create_index("projectRef")
        await db.calendar_events.create_index("startDate")
        await db.calendar_events.create_index("assignedTo", sparse=True)
        await db.calendar_events.create_index([("assignedTo", 1), ("startDate", 1)])
        await db.activities.create_index("assignedTo", sparse=True)
        await db.activities.create_index("date", sparse=True)
        await db.prescriptor_notes.create_index("prescriptorId", sparse=True)
        await db.prescriptor_notes.create_index([("prescriptorId", 1), ("date", 1)])
        await db.sale_fichas.create_index("id", unique=True, sparse=True)
        await db.sale_fichas.create_index("createdBy", sparse=True)
        await db.sale_fichas.create_index("createdAt")
        await db.sale_ficha_docs.create_index("fichaId")
        await db.google_calendar_connections.create_index("userId", unique=True, sparse=True)
        await db.google_calendar_links.create_index([("userId", 1), ("module", 1), ("erpId", 1)])
        logger.info("Índices adicionales de rendimiento creados")

        logger.info("✅ Todos los índices de base de datos configurados correctamente")
        
    except Exception as e:
        logger.warning(f"Error creando índices (algunos pueden ya existir): {e}")
    
    # =============================================
    # CREAR/ACTUALIZAR USUARIO MASTER
    # =============================================
    try:
        now = datetime.now(timezone.utc)
        # Campos de rol/permisos del usuario master (idempotentes en cada arranque).
        # IMPORTANTE: la contraseña NO se incluye aquí para no sobrescribirla en cada reinicio.
        master_fields = {
            "username": "MARIO",
            "email": "mario@luiggihome.es",
            "clientName": "ADMINISTRADOR MASTER",
            "phone": "",
            "province": "",
            "isAdmin": True,
            "isGerente": True,
            "isDirectorComercial": True,
            "isDirectorFabrica": True,
            "isResponsableDelegacion": True,
            "isComercial": False,
            "isTienda": False,
            "isPrescriptor": False,
            "canManageArticles": True,
            "canViewMetrics": True,
            "canExportData": True,
            "canManageUsers": True,
            "canManageClients": True,
            "canManageProducts": True,
            "canManageSettings": True,
            "canViewAllOrders": True,
            "canEditAllOrders": True,
            "canDeleteOrders": True,
            "canAccessTelemetry": True,
            "canAccessBackups": True,
            "canAccessMaintenance": True,
            "canAccessArmarios": True,
            "allowedModules": ["montada", "despiece", "armarios"],
            "active": True,
            "isActive": True,
            "updatedAt": now,
        }

        # La contraseña del master se toma de la variable de entorno MASTER_PASSWORD.
        # Si se define, se (re)establece en cada arranque (permite rotarla desde Railway).
        master_password = os.environ.get('MASTER_PASSWORD')
        existing_master = await db.users.find_one({"username": "MARIO"}, {"_id": 0, "password": 1})

        if not existing_master:
            # Primera creación: si no hay MASTER_PASSWORD, generar una temporal aleatoria
            # y registrarla UNA sola vez para no quedar bloqueados.
            if master_password:
                initial_password = master_password
                log_temp = False
            else:
                initial_password = secrets.token_urlsafe(12)
                log_temp = True
            doc = {
                **master_fields,
                "id": "user-master-mario",
                "password": bcrypt.hashpw(initial_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),
                "createdAt": now,
            }
            await db.users.insert_one(doc)
            if log_temp:
                logger.warning(
                    "⚠️ Usuario MASTER (MARIO) creado con contraseña TEMPORAL: %s — "
                    "cámbiala de inmediato y define MASTER_PASSWORD.", initial_password
                )
            else:
                logger.info("✅ Usuario MASTER (MARIO) creado.")
        else:
            update = {"$set": dict(master_fields)}
            if master_password:
                update["$set"]["password"] = bcrypt.hashpw(
                    master_password.encode('utf-8'), bcrypt.gensalt()
                ).decode('utf-8')
            await db.users.update_one({"username": "MARIO"}, update)
            logger.info("✅ Usuario MASTER (MARIO) actualizado (roles/permisos).")
    except Exception as e:
        logger.warning(f"Error configurando usuario master: {e}")
    
    # =============================================
    # INICIALIZAR SERVICIOS DE BACKUP Y TRACKING
    # =============================================
    try:
        # Inicializar servicio de backup con programador diario (3:00 AM)
        backup_svc = init_backup_service(mongo_url, os.environ['DB_NAME'])
        await backup_svc.start_scheduler(hour=3)
        logger.info("✅ Servicio de backup diario iniciado (3:00 AM)")
        
        # Inicializar tracker de actividad
        tracker = init_activity_tracker(db)
        await tracker.setup_indexes()
        logger.info("✅ Tracker de actividad de usuarios iniciado")
        
    except Exception as e:
        logger.warning(f"Error inicializando servicios de backup/tracking: {e}")
    
    # Start backup scheduler from backup module
    start_backup_scheduler()

    # Recordatorios de calendario por email (avisa al usuario asignado)
    try:
        from services.calendar_reminder_service import start_reminder_scheduler
        start_reminder_scheduler()
    except Exception as e:
        logger.warning(f"No se pudo iniciar el programador de recordatorios: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    backup_scheduler.shutdown()
    client.close()