# © 2024-2026 Luiggi Home. Todos los derechos reservados. [LUIGGI-COPYRIGHT]
# Software propietario y confidencial. Ver LICENSE.
# Prohibida su copia, distribución, modificación o uso sin autorización
# escrita del titular.
"""
plan_negocio_excel.py — EL PLAN, EN UNA HOJA DE CÁLCULO DE VERDAD.

CON FÓRMULAS, NO CON RESULTADOS PEGADOS
---------------------------------------
Fuera del ERP el plan tiene que seguir calculando: se toca el precio del casco y
cambia el margen, el punto de equilibrio y el EBITDA. Un Excel con los números
fijos es una FOTO, y una foto de un plan de negocio se queda vieja el mismo día
que alguien pregunta «¿y si el casco sube un 8 %?».

Por eso aquí se escriben las mismas cuentas que hace `services/plan_negocio.py`,
pero en el idioma de Excel. Las dos versiones tienen que decir lo mismo, y hay
un candado que lo comprueba con números.

LO QUE NO SE SABE SIGUE SIN SABERSE
-----------------------------------
Una casilla vacía sale VACÍA y en amarillo, y la fórmula que depende de ella
pone «falta dato». Nunca un cero: un cero parece un resultado.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FUENTE = "Arial"
AZUL = Font(name=FUENTE, size=10, color="0000FF")
NEGRO = Font(name=FUENTE, size=10)
VERDE = Font(name=FUENTE, size=10, color="008000")
TITULO = Font(name=FUENTE, size=14, bold=True, color="1F3864")
CAB = Font(name=FUENTE, size=10, bold=True, color="FFFFFF")
NOTA = Font(name=FUENTE, size=9, italic=True, color="7F7F7F")
FUERTE = Font(name=FUENTE, size=10, bold=True)

AMARILLO = PatternFill("solid", fgColor="FFFF00")
AZUL_CAB = PatternFill("solid", fgColor="1F3864")
CLARO = PatternFill("solid", fgColor="DDEBF7")

EUR = '#,##0.00 "€"'
EUR0 = '#,##0 "€"'
PCT = '0.0%'
NUM = '#,##0'
NUM1 = '#,##0.0'

_fino = Side(style="thin", color="BFBFBF")
BORDE = Border(left=_fino, right=_fino, top=_fino, bottom=_fino)

MATERIALES = ("casco", "bisagras", "guias", "patasZocalo", "otrosHerrajes",
              "componentes", "embalaje", "otrosDirectos")
ETIQUETA_MATERIAL = {
    "casco": "Casco (compra)", "bisagras": "Bisagras", "guias": "Guías",
    "patasZocalo": "Patas y zócalo", "otrosHerrajes": "Otros herrajes",
    "componentes": "Componentes", "embalaje": "Embalaje",
    "otrosDirectos": "Otros directos",
}
B2C = ("comisionComercial", "montaje", "transporte", "postventa")
ETIQUETA_B2C = {"comisionComercial": "Comisión comercial", "montaje": "Montaje",
                "transporte": "Transporte y logística", "postventa": "Postventa"}


def _valor(v):
    """Lo que se escribe en la casilla: el número, o nada.

    Nada es NADA, no un cero. Es toda la diferencia entre «esto está sin
    rellenar» y «esto vale cero euros».
    """
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _entrada(ws, celda, valor, formato=EUR):
    """Una casilla que rellena el master: azul si tiene dato, amarilla si no."""
    v = _valor(valor)
    ws[celda] = v
    ws[celda].number_format = formato
    ws[celda].border = BORDE
    if v is None:
        ws[celda].fill = AMARILLO
    ws[celda].font = AZUL
    return v


def _cab(ws, fila, valores):
    for i, v in enumerate(valores, start=1):
        c = ws.cell(row=fila, column=i, value=v)
        c.font = CAB
        c.fill = AZUL_CAB
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDE


def construir(datos):
    """Devuelve el libro de Excel del plan `datos`."""
    d = datos or {}
    wb = Workbook()

    _hoja_capacidad(wb, d)
    _hoja_coste(wb, d)
    _hoja_precios(wb, d)
    _hoja_bloque(wb, d, "Estructura", "GASTOS DE ESTRUCTURA (AL AÑO)", "estructura")
    _hoja_bloque(wb, d, "Inversion", "INVERSIÓN INICIAL", "inversion")
    _hoja_equilibrio(wb, d)

    del wb["Sheet"]
    wb._sheets.sort(key=lambda s: ["Capacidad", "Coste_mueble", "Precios",
                                   "Estructura", "Inversion", "Equilibrio"].index(s.title))
    return wb


# ─── Capacidad ──────────────────────────────────────────────────────────────

def _hoja_capacidad(wb, d):
    ws = wb.create_sheet("Capacidad")
    ws.sheet_view.showGridLines = False
    for col, ancho in {"A": 42, "B": 16, "C": 14, "D": 56}.items():
        ws.column_dimensions[col].width = ancho
    ws["A1"] = "1 · CAPACIDAD DE FÁBRICA"
    ws["A1"].font = TITULO
    _cab(ws, 3, ["Concepto", "Valor", "Unidad", "Nota"])

    cap = d.get("capacidad") or {}
    filas = [
        ("Personas en fabricación", "personas", NUM, "personas", ""),
        ("Productividad del equipo", "mueblesHora", NUM1, "muebles/hora",
         "CONJUNTA de todo el equipo, no por persona."),
        ("Jornada", "horasDia", NUM1, "horas/día", ""),
        ("Días laborables", "diasSemana", NUM, "días/semana", ""),
        ("Horas productivas al año", "horasAno", NUM, "horas/año",
         "Del EQUIPO, ya sin vacaciones ni festivos."),
    ]
    for i, (nombre, clave, fmt, unidad, nota) in enumerate(filas):
        f = 4 + i
        ws.cell(row=f, column=1, value=nombre).font = NEGRO
        _entrada(ws, f"B{f}", cap.get(clave), fmt)
        ws.cell(row=f, column=3, value=unidad).font = NOTA
        ws.cell(row=f, column=4, value=nota).font = NOTA

    ws["A9"] = "CAPACIDAD DE REFERENCIA"
    ws["A9"].font = FUERTE
    ws["B9"] = '=IF(OR($B$5="",$B$8=""),"falta dato",$B$5*$B$8)'
    ws["B9"].font = NEGRO
    ws["B9"].number_format = NUM
    ws["B9"].fill = CLARO
    ws["C9"] = "muebles/año"
    ws["C9"].font = NOTA
    ws["D9"] = "Es un TECHO, no un objetivo de venta."
    ws["D9"].font = NOTA

    ws["A10"] = "Capacidad semanal"
    ws["A10"].font = NEGRO
    ws["B10"] = '=IF(OR($B$5="",$B$6="",$B$7=""),"falta dato",$B$5*$B$6*$B$7)'
    ws["B10"].font = NEGRO
    ws["B10"].number_format = NUM
    ws["C10"] = "muebles/semana"
    ws["C10"].font = NOTA

    ws["A12"] = "Coste empresa por persona y hora"
    ws["A12"].font = NEGRO
    _entrada(ws, "B12", cap.get("costeHoraPersona"), EUR)
    ws["C12"] = "€/hora"
    ws["C12"].font = NOTA
    ws["D12"] = "Bruto + Seguridad Social de la empresa. No es el sueldo neto."
    ws["D12"].font = NOTA

    ws["A13"] = "MANO DE OBRA POR MUEBLE"
    ws["A13"].font = FUERTE
    ws["B13"] = '=IF(OR($B$12="",$B$4="",$B$5=""),"falta dato",$B$4*$B$12/$B$5)'
    ws["B13"].font = NEGRO
    ws["B13"].number_format = EUR
    ws["B13"].fill = CLARO
    ws["C13"] = "€/mueble"
    ws["C13"].font = NOTA
    ws["D13"] = "Coste del equipo por hora entre los muebles de esa hora."
    ws["D13"].font = NOTA


# ─── Coste por mueble ───────────────────────────────────────────────────────

def _hoja_coste(wb, d):
    ws = wb.create_sheet("Coste_mueble")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "2 · COSTE REAL POR MUEBLE"
    ws["A1"].font = TITULO
    ws["A2"] = "Sin el precio de compra del casco no hay plan: es el dato que desbloquea todo lo demás."
    ws["A2"].font = NOTA

    cols = ["Referencia"] + [ETIQUETA_MATERIAL[c] for c in MATERIALES] + \
           ["MATERIALES", "Mano de obra", "COSTE DIRECTO"]
    for i, ancho in enumerate([15] + [12] * len(MATERIALES) + [13, 13, 14]):
        ws.column_dimensions[get_column_letter(1 + i)].width = ancho
    _cab(ws, 4, cols)

    refs = d.get("referencias") or []
    n = len(refs)
    col_mat = 2 + len(MATERIALES)          # J si hay 8 materiales
    col_mo = col_mat + 1
    col_tot = col_mo + 1
    for i, r in enumerate(refs):
        f = 5 + i
        ws.cell(row=f, column=1, value=r.get("nombre") or "").font = FUERTE
        ws.cell(row=f, column=1).border = BORDE
        for j, clave in enumerate(MATERIALES):
            _entrada(ws, f"{get_column_letter(2 + j)}{f}", r.get(clave), EUR)
        rango = f"{get_column_letter(2)}{f}:{get_column_letter(1 + len(MATERIALES))}{f}"
        # Si falta una sola partida, NO se suma lo que hay: un coste a medias no
        # es un coste bajo, es un coste desconocido.
        c = ws.cell(row=f, column=col_mat,
                    value=f'=IF(COUNT({rango})<{len(MATERIALES)},"falta dato",SUM({rango}))')
        c.font = NEGRO
        c.number_format = EUR
        c.border = BORDE
        mo = ws.cell(row=f, column=col_mo, value="=Capacidad!$B$13")
        mo.font = VERDE
        mo.number_format = EUR
        mo.border = BORDE
        L = get_column_letter(col_mat)
        M = get_column_letter(col_mo)
        t = ws.cell(row=f, column=col_tot,
                    value=f'=IF(OR(NOT(ISNUMBER({L}{f})),NOT(ISNUMBER({M}{f}))),"falta dato",{L}{f}+{M}{f})')
        t.font = FUERTE
        t.number_format = EUR
        t.fill = CLARO
        t.border = BORDE

    f = 5 + n + 1
    ws.cell(row=f, column=1, value="Coste directo medio").font = FUERTE
    T = get_column_letter(col_tot)
    c = ws.cell(row=f, column=col_tot,
                value=f'=IF(COUNT({T}5:{T}{4 + n})=0,"falta dato",AVERAGE({T}5:{T}{4 + n}))')
    c.font = NEGRO
    c.number_format = EUR
    c.fill = CLARO


# ─── Precios y margen ───────────────────────────────────────────────────────

def _hoja_precios(wb, d):
    ws = wb.create_sheet("Precios")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "3 · PRECIO DE VENTA Y MARGEN"
    ws["A1"].font = TITULO
    ws["A2"] = "El margen se calcula sobre el PRECIO DE VENTA, no sobre el coste: 100 de coste y 150 de venta son un 33 %, no un 50 %."
    ws["A2"].font = NOTA

    for i, ancho in enumerate([15, 14, 14, 14, 12, 14, 14, 12]):
        ws.column_dimensions[get_column_letter(1 + i)].width = ancho
    _cab(ws, 4, ["Referencia", "Coste directo", "Precio B2B", "Margen B2B €",
                 "Margen B2B %", "Precio B2C", "Margen B2C €", "Margen B2C %"])

    refs = d.get("referencias") or []
    n = len(refs)
    col_tot = 2 + len(MATERIALES) + 2      # columna COSTE DIRECTO en Coste_mueble
    T = get_column_letter(col_tot)
    for i, r in enumerate(refs):
        f = 5 + i
        fc = 5 + i
        ws.cell(row=f, column=1, value=r.get("nombre") or "").font = FUERTE
        c = ws.cell(row=f, column=2, value=f"=Coste_mueble!{T}{fc}")
        c.font = VERDE
        c.number_format = EUR
        c.border = BORDE
        _entrada(ws, f"C{f}", r.get("precioB2B"), EUR)
        ws.cell(row=f, column=4, value=f'=IF(OR(C{f}="",NOT(ISNUMBER(B{f}))),"falta dato",C{f}-B{f})').font = NEGRO
        ws.cell(row=f, column=4).number_format = EUR
        ws.cell(row=f, column=5, value=f'=IF(OR(C{f}="",C{f}=0,NOT(ISNUMBER(B{f}))),"falta dato",(C{f}-B{f})/C{f})').font = NEGRO
        ws.cell(row=f, column=5).number_format = PCT
        _entrada(ws, f"F{f}", r.get("precioB2C"), EUR)
        ws.cell(row=f, column=7, value=f'=IF(OR(F{f}="",NOT(ISNUMBER(B{f}))),"falta dato",F{f}-B{f})').font = NEGRO
        ws.cell(row=f, column=7).number_format = EUR
        ws.cell(row=f, column=8, value=f'=IF(OR(F{f}="",F{f}=0,NOT(ISNUMBER(B{f}))),"falta dato",(F{f}-B{f})/F{f})').font = NEGRO
        ws.cell(row=f, column=8).number_format = PCT

    fm = 5 + n + 1
    ws.cell(row=fm, column=1, value="MEDIA").font = FUERTE
    for col in ("C", "D", "F", "G"):
        c = ws[f"{col}{fm}"]
        c.value = f'=IF(COUNT({col}5:{col}{4 + n})=0,"falta dato",AVERAGE({col}5:{col}{4 + n}))'
        c.font = NEGRO
        c.number_format = EUR
        c.fill = CLARO

    f0 = fm + 2
    ws.cell(row=f0, column=1, value="Costes del B2C que no existen en B2B").font = FUERTE
    ws.cell(row=f0 + 1, column=4,
            value="Aquí un 0 SÍ vale: no pagar comisión es un dato.").font = NOTA
    b2c = d.get("b2c") or {}
    for i, clave in enumerate(B2C):
        f = f0 + 1 + i
        ws.cell(row=f, column=1, value=ETIQUETA_B2C[clave]).font = NEGRO
        _entrada(ws, f"B{f}", b2c.get(clave), EUR)
    ftot = f0 + 1 + len(B2C)
    ws.cell(row=ftot, column=1, value="Total costes B2C").font = FUERTE
    ws[f"B{ftot}"] = f'=IF(COUNT(B{f0 + 1}:B{ftot - 1})<{len(B2C)},"falta dato",SUM(B{f0 + 1}:B{ftot - 1}))'
    ws[f"B{ftot}"].font = NEGRO
    ws[f"B{ftot}"].number_format = EUR

    fmc = ftot + 1
    ws.cell(row=fmc, column=1, value="MARGEN DE CONTRIBUCIÓN B2C").font = FUERTE
    ws[f"B{fmc}"] = f'=IF(OR(NOT(ISNUMBER($G${fm})),NOT(ISNUMBER($B${ftot}))),"falta dato",$G${fm}-$B${ftot})'
    ws[f"B{fmc}"].font = NEGRO
    ws[f"B{fmc}"].number_format = EUR
    ws[f"B{fmc}"].fill = CLARO

    frep = fmc + 2
    ws.cell(row=frep, column=1, value="% de ventas en B2B (0,7 = 70 %)").font = NEGRO
    _entrada(ws, f"B{frep}", d.get("repartoB2B"), PCT)
    fmm = frep + 1
    ws.cell(row=fmm, column=1, value="MARGEN MEDIO PONDERADO POR MUEBLE").font = FUERTE
    ws[f"B{fmm}"] = (f'=IF(OR($B${frep}="",NOT(ISNUMBER($D${fm})),NOT(ISNUMBER($B${fmc}))),"falta dato",'
                     f'$B${frep}*$D${fm}+(1-$B${frep})*$B${fmc})')
    ws[f"B{fmm}"].font = NEGRO
    ws[f"B{fmm}"].number_format = EUR
    ws[f"B{fmm}"].fill = CLARO

    fpm = fmm + 1
    ws.cell(row=fpm, column=1, value="Precio medio ponderado").font = NEGRO
    ws[f"B{fpm}"] = (f'=IF(OR($B${frep}="",NOT(ISNUMBER($C${fm})),NOT(ISNUMBER($F${fm}))),"falta dato",'
                     f'$B${frep}*$C${fm}+(1-$B${frep})*$F${fm})')
    ws[f"B{fpm}"].font = NEGRO
    ws[f"B{fpm}"].number_format = EUR

    ws._alemar = {"margen": f"B{fmm}", "precio": f"B{fpm}"}


# ─── Bloques de importes (estructura, inversión) ────────────────────────────

def _hoja_bloque(wb, d, nombre, titulo, clave):
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws["A1"] = titulo
    ws["A1"].font = TITULO
    _cab(ws, 3, ["Concepto", "€"])

    bloque = d.get(clave) or {}
    f = 4
    for k, v in bloque.items():
        ws.cell(row=f, column=1, value=k).font = NEGRO
        _entrada(ws, f"B{f}", v, EUR0)
        f += 1
    ws.cell(row=f + 1, column=1, value=f"TOTAL {nombre.upper()}").font = FUERTE
    ws[f"B{f + 1}"] = (f'=IF(COUNT(B4:B{f - 1})<{max(1, f - 4)},"falta dato",SUM(B4:B{f - 1}))'
                       if f > 4 else '"falta dato"')
    ws[f"B{f + 1}"].font = FUERTE
    ws[f"B{f + 1}"].number_format = EUR0
    ws[f"B{f + 1}"].fill = CLARO
    ws._alemar_total = f"B{f + 1}"


# ─── Equilibrio ─────────────────────────────────────────────────────────────

def _hoja_equilibrio(wb, d):
    ws = wb.create_sheet("Equilibrio")
    ws.sheet_view.showGridLines = False
    for col, ancho in {"A": 46, "B": 18, "C": 15, "D": 52}.items():
        ws.column_dimensions[col].width = ancho
    ws["A1"] = "4 · CUÁNTOS PEDIDOS HACEN FALTA, Y CUÁNDO CONTRATAR"
    ws["A1"].font = TITULO
    _cab(ws, 3, ["Concepto", "Valor", "Unidad", "Nota"])

    margen = wb["Precios"]._alemar["margen"]
    precio = wb["Precios"]._alemar["precio"]
    estructura = wb["Estructura"]._alemar_total

    ws["A4"] = "Margen medio por mueble"
    ws["A4"].font = NEGRO
    ws["B4"] = f"=Precios!${margen[0]}${margen[1:]}"
    ws["B4"].font = VERDE
    ws["B4"].number_format = EUR
    ws["C4"] = "€/mueble"
    ws["C4"].font = NOTA

    ws["A5"] = "Gastos de estructura al año"
    ws["A5"].font = NEGRO
    ws["B5"] = f"=Estructura!${estructura[0]}${estructura[1:]}"
    ws["B5"].font = VERDE
    ws["B5"].number_format = EUR0

    ws["A6"] = "MUEBLES NECESARIOS PARA CUBRIR LA ESTRUCTURA"
    ws["A6"].font = FUERTE
    # Con margen 0 o negativo NO hay punto de equilibrio: no es un número muy
    # grande, es que por muchos muebles que se vendan no se cubre nada.
    ws["B6"] = '=IF(OR(NOT(ISNUMBER($B$4)),$B$4<=0,NOT(ISNUMBER($B$5))),"no existe: el margen no cubre",$B$5/$B$4)'
    ws["B6"].font = NEGRO
    ws["B6"].number_format = NUM
    ws["B6"].fill = CLARO
    ws["C6"] = "muebles/año"
    ws["C6"].font = NOTA

    ws["A7"] = "Capacidad de la fábrica"
    ws["A7"].font = NEGRO
    ws["B7"] = "=Capacidad!$B$9"
    ws["B7"].font = VERDE
    ws["B7"].number_format = NUM

    ws["A8"] = "Qué parte de la capacidad hay que vender para no perder"
    ws["A8"].font = NEGRO
    ws["B8"] = '=IF(OR(NOT(ISNUMBER($B$6)),NOT(ISNUMBER($B$7)),$B$7=0),"falta dato",$B$6/$B$7)'
    ws["B8"].font = NEGRO
    ws["B8"].number_format = PCT
    ws["D8"] = "Por encima del 100 % no sale con esta estructura: hay que bajar coste o subir precio."
    ws["D8"].font = NOTA

    ws["A10"] = "Facturación a plena capacidad"
    ws["A10"].font = NEGRO
    ws["B10"] = f'=IF(OR(NOT(ISNUMBER(Precios!${precio[0]}${precio[1:]})),NOT(ISNUMBER($B$7))),"falta dato",Precios!${precio[0]}${precio[1:]}*$B$7)'
    ws["B10"].font = NEGRO
    ws["B10"].number_format = EUR0

    ws["A11"] = "EBITDA a plena capacidad"
    ws["A11"].font = FUERTE
    ws["B11"] = '=IF(OR(NOT(ISNUMBER($B$4)),NOT(ISNUMBER($B$5)),NOT(ISNUMBER($B$7))),"falta dato",$B$4*$B$7-$B$5)'
    ws["B11"].font = NEGRO
    ws["B11"].number_format = EUR0
    ws["B11"].fill = CLARO

    ws["A13"] = "Plazo de entrega que se quiere prometer"
    ws["A13"].font = NEGRO
    _entrada(ws, "B13", d.get("plazoEntregaSemanas"), NUM)
    ws["C13"] = "semanas"
    ws["C13"].font = NOTA

    ws["A14"] = "CARTERA A PARTIR DE LA CUAL HAY QUE CONTRATAR"
    ws["A14"].font = FUERTE
    ws["B14"] = '=IF(OR($B$13="",NOT(ISNUMBER(Capacidad!$B$10))),"falta dato",$B$13*Capacidad!$B$10)'
    ws["B14"].font = NEGRO
    ws["B14"].number_format = NUM
    ws["B14"].fill = CLARO
    ws["C14"] = "muebles pendientes"
    ws["C14"].font = NOTA
    ws["D14"] = "Sostenido varias semanas, no una punta. Y comprobando que el margen aguanta."
    ws["D14"].font = NOTA
