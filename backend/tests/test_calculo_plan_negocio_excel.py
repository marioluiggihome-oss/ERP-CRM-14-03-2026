# © 2024-2026 Luiggi Home. Todos los derechos reservados. [LUIGGI-COPYRIGHT]
# Software propietario y confidencial. Ver LICENSE.
# Prohibida su copia, distribución, modificación o uso sin autorización
# escrita del titular.
"""CANDADO del Excel del plan: QUE DIGA LO MISMO QUE EL ERP.

Hay dos sitios donde se calcula el plan de negocio: dentro del ERP
(`services/plan_negocio.py`) y en el Excel que se descarga
(`services/plan_negocio_excel.py`). Es una duplicidad ACEPTADA a proposito:
fuera del ERP el plan tiene que seguir calculando cuando alguien toque el precio
del casco, y un Excel con los numeros pegados es una foto que se queda vieja el
mismo dia.

Pero una duplicidad aceptada hay que vigilarla. Que las dos versiones digan cosas
distintas es peor que tener una sola, porque entonces hay dos planes de negocio
y nadie sabe cual es el bueno.

LO QUE SE PROTEGE
-----------------
1. LAS FORMULAS APUNTAN DONDE TIENEN QUE APUNTAR. Una referencia a la fila de al
   lado no da error: da un plan con los numeros cruzados y buena pinta.

2. LAS REGLAS DE LA CASA ESTAN TAMBIEN EN EL EXCEL:
   · no se suma un coste a medias,
   · el margen va sobre el precio de venta,
   · con margen cero o negativo NO hay punto de equilibrio,
   · lo que falta se dice, no se pone a cero.

3. LO QUE ESTA SIN RELLENAR SALE VACIO Y EN AMARILLO, nunca a cero.
"""
import importlib.util
import io
import os

import pytest

BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _cargar(nombre):
    ruta = os.path.join(BACKEND, "services", f"{nombre}.py")
    spec = importlib.util.spec_from_file_location(f"_{nombre}", ruta)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


@pytest.fixture(scope="module")
def motor():
    return _cargar("plan_negocio")


@pytest.fixture(scope="module")
def xls():
    return _cargar("plan_negocio_excel")


@pytest.fixture(scope="module")
def datos(motor):
    """La fabrica del master, con importes de prueba para poder comprobar."""
    return {
        "capacidad": {"personas": 2, "mueblesHora": 3, "horasDia": 7,
                      "diasSemana": 5, "horasAno": 1600, "costeHoraPersona": 18},
        "referencias": [
            {"nombre": n, "casco": 40, "bisagras": 3, "guias": 8, "patasZocalo": 4,
             "otrosHerrajes": 2, "componentes": 6, "embalaje": 2, "otrosDirectos": 1,
             "precioB2B": 120, "precioB2C": 180}
            for n in motor.REFERENCIAS
        ],
        "b2c": {c: 0 for c in motor.COSTES_B2C},
        "repartoB2B": 1,
        "estructura": {"Alquiler": 24000, "Salarios": 60000},
        "inversion": {"Maquinaria": 60000},
        "plazoEntregaSemanas": 4,
    }


@pytest.fixture(scope="module")
def libro(xls, datos):
    return xls.construir(datos)


def _columna(ws, cabecera, fila_cabecera=4):
    """La letra de la columna que se llama asi.

    Se busca POR NOMBRE y no por letra a proposito: aniadir una columna delante
    corre todas las demas, y una prueba atada a la letra empieza a comprobar la
    columna de al lado sin enterarse — que es peor que no comprobar nada.
    """
    from openpyxl.utils import get_column_letter
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=fila_cabecera, column=c).value == cabecera:
            return get_column_letter(c)
    raise AssertionError(f"no hay columna «{cabecera}» en {ws.title}")


# ─── 1. Las formulas apuntan donde tienen que apuntar ───────────────────────

def test_cada_referencia_de_precios_mira_SU_coste(libro, motor):
    """CANDADO. Un desfase de una fila no da error: da el coste del Bajo 45
    puesto contra el precio del Bajo 60, y un margen que no existe."""
    p, cm = libro["Precios"], libro["Coste_mueble"]
    for i in range(len(motor.REFERENCIAS)):
        f = 5 + i
        nombre_precios = p.cell(row=f, column=1).value
        formula = p.cell(row=f, column=2).value
        assert formula.startswith("=Coste_mueble!")
        fila_cm = int("".join(ch for ch in formula.split("!")[1] if ch.isdigit()))
        assert cm.cell(row=fila_cm, column=1).value == nombre_precios, (
            f"la fila de {nombre_precios} en Precios mira el coste de "
            f"{cm.cell(row=fila_cm, column=1).value}")


def test_la_columna_que_mira_es_la_del_COSTE_DIRECTO(libro):
    """Mirar la de material dejaria la mano de obra fuera del margen."""
    p, cm = libro["Precios"], libro["Coste_mueble"]
    col = p["B5"].value.split("!")[1].rstrip("0123456789")
    assert cm[f"{col}4"].value == "COSTE DIRECTO"
    assert col == _columna(cm, "COSTE DIRECTO")


def test_el_equilibrio_mira_el_margen_ponderado_y_la_estructura(libro):
    e, p, s = libro["Equilibrio"], libro["Precios"], libro["Estructura"]

    def _destino(formula, hoja):
        ref = formula.split("!")[1].replace("$", "")
        return hoja[ref.rstrip("0123456789") + "".join(c for c in ref if c.isdigit())]

    fila_margen = int("".join(c for c in e["B4"].value.split("!")[1] if c.isdigit()))
    assert p.cell(row=fila_margen, column=1).value == "MARGEN MEDIO PONDERADO POR MUEBLE"
    fila_est = int("".join(c for c in e["B5"].value.split("!")[1] if c.isdigit()))
    assert s.cell(row=fila_est, column=1).value == "TOTAL ESTRUCTURA"


def test_la_mano_de_obra_sale_del_TIEMPO_de_cada_mueble(libro):
    """CANDADO. Es la unidad que se mide: la hora de fabrica, repartida por los
    minutos que lleva ese mueble. Sin medir, la media — y eso es un dato peor,
    porque da lo mismo a un alto de 45 que a un bajo de 90 con cajones."""
    cm, cap = libro["Coste_mueble"], libro["Capacidad"]
    col_min = _columna(cm, "Min / mueble")
    formula = cm[f"{_columna(cm, 'Mano de obra')}5"].value
    assert f'IF({col_min}5<>""' in formula, "el tiempo medido no manda"
    assert "/60" in formula, "los minutos tienen que pasarse a horas"
    fila_hora = int(formula.split("Capacidad!$B$")[1].split("*")[0])
    assert cap.cell(row=fila_hora, column=1).value == "HORA DE FÁBRICA (equipo)"
    fila_media = int(formula.rsplit("Capacidad!$B$", 1)[1].rstrip(")"))
    assert cap.cell(row=fila_media, column=1).value == "Mano de obra por mueble (media)"


def test_ninguna_formula_apunta_a_una_hoja_que_no_existe(libro):
    import re
    hojas = set(libro.sheetnames)
    for ws in libro:
        for fila in ws.iter_rows():
            for c in fila:
                if isinstance(c.value, str) and c.value.startswith("="):
                    for m in re.finditer(r"([A-Za-z_][A-Za-z0-9_]*)!", c.value):
                        assert m.group(1) in hojas, f"{ws.title}!{c.coordinate} → {m.group(1)}"


# ─── 2. Las reglas de la casa, tambien en el Excel ──────────────────────────

def test_el_excel_tampoco_suma_un_coste_a_medias(libro, motor):
    """Si falta una partida, el Excel tiene que decir «falta dato», no sumar lo
    que hay: un coste a medias da un margen enorme y un plan precioso."""
    cm = libro["Coste_mueble"]
    formula = cm[f"{_columna(cm, 'MATERIAL')}5"].value
    assert "COUNT(" in formula and "falta dato" in formula
    assert f"<{len(motor.COSTES_DE_MATERIAL)}" in formula


def test_el_coste_por_mueble_MANDA_sobre_el_desglose_tambien_en_el_excel(libro):
    """CANDADO. Es la misma regla que dentro del ERP. Si el Excel sumara los dos
    contaria el material dos veces, y habria dos planes que no coinciden."""
    cm = libro["Coste_mueble"]
    col_mueble = _columna(cm, "Material / mueble")
    formula = cm[f"{_columna(cm, 'MATERIAL')}5"].value
    assert formula.startswith(f'=IF({col_mueble}5<>""'), (
        "el Excel no da prioridad al coste escrito a mano por mueble: "
        f"«{formula}»")


def test_la_columna_por_mueble_va_la_PRIMERA(libro):
    """Es como se empieza. Si estuviera detras de ocho columnas de desglose,
    nadie la encontraria y todos rellenarian lo dificil."""
    cm = libro["Coste_mueble"]
    assert _columna(cm, "Material / mueble") == "B"


def test_el_margen_del_excel_va_sobre_el_precio_de_venta(libro):
    """(precio − coste) / PRECIO. Dividir por el coste da un 50 % donde hay 33."""
    formula = libro["Precios"]["E5"].value
    assert "(C5-B5)/C5" in formula.replace(" ", "")


def test_con_margen_cero_o_negativo_no_hay_punto_de_equilibrio(libro):
    """No es un numero grandisimo de muebles: es que no existe. Vendiendo por
    debajo del coste, no hay cantidad que cubra la estructura."""
    formula = libro["Equilibrio"]["B6"].value
    assert "$B$4<=0" in formula
    assert "no existe" in formula


def test_el_excel_avisa_de_que_por_encima_del_100_no_sale(libro):
    assert "100" in (libro["Equilibrio"]["D8"].value or "")


# ─── 3. Lo que falta sale vacio y en amarillo ───────────────────────────────

def test_una_casilla_sin_rellenar_sale_VACIA_no_a_cero(xls, motor):
    """CANDADO. Un casco a 0 € en el Excel da un margen del 100 %, igual que en
    el ERP. Vacia y amarilla, que se vea."""
    datos = {
        "capacidad": {"personas": 2, "mueblesHora": 3, "horasAno": 1600},
        "referencias": [{"nombre": "Bajo 60"}],
        "estructura": {}, "inversion": {},
    }
    wb = xls.construir(datos)
    cm = wb["Coste_mueble"]
    for cabecera in ("Material / mueble", "Casco (compra)"):
        c = cm[f"{_columna(cm, cabecera)}5"]
        assert c.value is None, f"«{cabecera}» sin dato NO puede salir a cero"
        assert c.fill.fgColor.rgb.endswith("FFFF00"), "y tiene que verse en amarillo"
    assert wb["Capacidad"]["B12"].value is None


def test_lo_que_si_se_sabe_sale_escrito(libro):
    assert libro["Capacidad"]["B4"].value == 2      # personas
    assert libro["Capacidad"]["B5"].value == 3      # muebles/hora
    assert libro["Capacidad"]["B8"].value == 1600   # horas/año
    cm = libro["Coste_mueble"]
    assert cm[f"{_columna(cm, 'Casco (compra)')}5"].value == 40


def test_las_seis_referencias_estan(libro, motor):
    nombres = [libro["Coste_mueble"].cell(row=5 + i, column=1).value
               for i in range(len(motor.REFERENCIAS))]
    assert nombres == list(motor.REFERENCIAS)


# ─── 4. Que el fichero se pueda abrir de verdad ─────────────────────────────

def test_el_libro_se_guarda_y_se_vuelve_a_leer(libro):
    """Un generador que produce un fichero que Excel no abre no sirve de nada, y
    eso no se ve hasta que alguien intenta abrirlo."""
    from openpyxl import load_workbook
    buf = io.BytesIO()
    libro.save(buf)
    buf.seek(0)
    wb2 = load_workbook(buf)
    assert wb2.sheetnames == ["Capacidad", "Coste_mueble", "Precios",
                              "Estructura", "Inversion", "Equilibrio"]


def test_no_se_usan_funciones_que_libreoffice_no_entiende(libro):
    """XLOOKUP y compania quedan como #NAME? en cuanto el fichero se abre en
    algo que no sea el Excel mas nuevo."""
    prohibidas = ("XLOOKUP", "XMATCH", "TEXTJOIN", "IFS(", "FILTER(", "UNIQUE(")
    for ws in libro:
        for fila in ws.iter_rows():
            for c in fila:
                if isinstance(c.value, str) and c.value.startswith("="):
                    for p in prohibidas:
                        assert p not in c.value.upper(), f"{ws.title}!{c.coordinate}: {p}"
